from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
import jwt
from passlib.context import CryptContext
import io
from openpyxl import Workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'fpo_database')]

# Create the main app
app = FastAPI(title="FPO Management System API")

# Health check endpoint at root level (required for Kubernetes deployment)
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes probes"""
    return {"status": "healthy", "service": "fpo-backend"}

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET', 'fpo-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ===================== MODELS =====================

class UserBase(BaseModel):
    username: str
    full_name: str
    role: str = "farmer"  # admin, agent, or farmer
    outlet_id: Optional[str] = None
    mobile: Optional[str] = None
    village: Optional[str] = None
    is_active: bool = True
    status: str = "active"  # active, pending, rejected (for agent approval)

class UserCreate(UserBase):
    password: str

class User(UserBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class UserInDB(User):
    hashed_password: str

class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    full_name: str
    role: str = "farmer"  # farmer or agent (admin can only be created by admin)
    mobile: Optional[str] = None
    village: Optional[str] = None

class ApproveAgentRequest(BaseModel):
    user_id: str
    outlet_id: str
    approved: bool = True

class PasswordChangeRequest(BaseModel):
    old_password: str
    new_password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

# Buy/Sell Request from Farmers
class ProductRequestBase(BaseModel):
    product_id: str
    quantity: float
    request_type: str  # buy or sell
    preferred_rate: Optional[float] = None
    notes: Optional[str] = None
    outlet_id: Optional[str] = None
    custom_product_name: Optional[str] = None  # For sell requests with custom products

class ProductRequest(ProductRequestBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    farmer_id: str
    farmer_name: Optional[str] = None
    product_name: Optional[str] = None
    status: str = "pending"  # pending, approved, rejected, completed
    outlet_id: Optional[str] = None  # assigned outlet for processing
    processed_by: Optional[str] = None
    processed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class OutletBase(BaseModel):
    name: str
    address: str
    contact_person: str
    mobile: Optional[str] = None
    is_central: bool = False
    is_active: bool = True

class OutletCreate(OutletBase):
    pass

class Outlet(OutletBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class ProductVariety(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    name_hi: Optional[str] = None

class ProductBase(BaseModel):
    name: str
    name_hi: Optional[str] = None  # Hindi name
    unit: str = "kg"  # kg, litre, piece, packet
    category: str = "input"  # input (seeds, bio-inputs) or output (produce)
    description: Optional[str] = None
    is_active: bool = True
    varieties: List[ProductVariety] = []  # Optional product varieties (e.g., Basmati, Sona Masoori)

class ProductCreate(ProductBase):
    pass

class Product(ProductBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class StockBase(BaseModel):
    product_id: str
    outlet_id: str
    quantity: float = 0
    opening_stock: float = 0
    stock_received: float = 0
    stock_sold: float = 0
    stock_damaged: float = 0

class Stock(StockBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class StockMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    from_outlet_id: Optional[str] = None  # None for new stock entry
    to_outlet_id: Optional[str] = None  # None for damage/discard
    quantity: float
    movement_type: str  # allocation, damage, adjustment, purchase, sale
    reason: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=datetime.utcnow)

class StockAllocationRequest(BaseModel):
    product_id: str
    from_outlet_id: str
    to_outlet_id: str
    quantity: float

class StockDamageRequest(BaseModel):
    product_id: str
    outlet_id: str
    quantity: float
    reason: str  # Mandatory reason

class StockAddRequest(BaseModel):
    product_id: str
    outlet_id: str
    quantity: float

# Stock Transfer Request System
class StockTransferRequestCreate(BaseModel):
    product_id: str
    from_outlet_id: str
    to_outlet_id: str
    quantity: float
    reason: Optional[str] = None

class StockTransferRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    from_outlet_id: str
    from_outlet_name: str
    to_outlet_id: str
    to_outlet_name: str
    quantity: float
    reason: Optional[str] = None
    status: str = "pending"  # pending, approved, rejected
    requested_by: str
    requested_by_name: str
    approved_by: Optional[str] = None
    approved_by_name: Optional[str] = None
    admin_remark: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class CustomerBase(BaseModel):
    name: str
    name_hi: Optional[str] = None  # Hindi name
    mobile: Optional[str] = None
    address: Optional[str] = None
    village: Optional[str] = None
    customer_type: str = "walk_in"  # walk_in, registered, shareholder
    linked_farmer_id: Optional[str] = None  # Link to farmer if customer is a farmer
    folio_number: Optional[str] = None  # For shareholders
    is_active: bool = True

class CustomerCreate(CustomerBase):
    pass

class Customer(CustomerBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    total_purchases: float = 0
    total_credit: float = 0
    total_paid: float = 0
    outstanding_balance: float = 0
    transaction_count: int = 0  # Total number of transactions
    last_transaction_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class SaleItemBase(BaseModel):
    product_id: str
    product_name: str
    quantity: float
    rate: float
    amount: float
    variety_id: Optional[str] = None  # Optional product variety
    variety_name: Optional[str] = None

class SaleBase(BaseModel):
    outlet_id: str
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_mobile: Optional[str] = None  # Optional mobile number
    items: List[SaleItemBase]
    subtotal: float
    discount: float = 0
    total_amount: float
    payment_mode: str  # cash, online, credit, partial
    cash_amount: float = 0
    online_amount: float = 0
    credit_amount: float = 0
    notes: Optional[str] = None
    # Shareholder info for discount application
    is_shareholder: bool = False
    shareholder_discount: float = 0

class SaleCreate(SaleBase):
    pass

class Sale(SaleBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_number: str = ""
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)
    synced: bool = True
    # Deletion tracking
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None
    deletion_reason: Optional[str] = None

# Transaction Deletion Audit Log
class TransactionDeletionLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_type: str  # "sale", "vendor_procurement", "farmer_purchase"
    transaction_id: str
    original_data: dict  # Store the original transaction data
    bill_number: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = None
    total_amount: float
    deleted_by: str  # Admin user who deleted
    deleted_by_name: str
    deletion_reason: Optional[str] = None
    reversal_details: dict  # Details of what was reversed
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Outlet Cash Deposit (Agent submits cash to Admin)
class OutletCashDepositCreate(BaseModel):
    outlet_id: str
    amount: float
    payment_mode: str = "cash"  # cash, online, bank_transfer
    reference_number: Optional[str] = None  # For bank/online transfers
    notes: Optional[str] = None

class OutletCashDeposit(OutletCashDepositCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    outlet_name: Optional[str] = None
    deposited_by: str = ""  # Agent user ID
    deposited_by_name: Optional[str] = None
    received_by: Optional[str] = None  # Admin who received (for confirmation)
    received_by_name: Optional[str] = None
    status: str = "pending"  # pending, confirmed, rejected
    confirmed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class FarmerBase(BaseModel):
    name: str
    village: str
    mobile: Optional[str] = None
    is_member: bool = False
    is_shareholder: bool = False
    folio_number: Optional[str] = None  # Share certificate folio number
    share_value: Optional[float] = None  # Value of shares held
    share_certificate_url: Optional[str] = None
    is_active: bool = True

class FarmerCreate(FarmerBase):
    pass

class Farmer(FarmerBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    total_supplied: float = 0
    total_payable: float = 0
    total_paid: float = 0
    outstanding_dues: float = 0
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

# Vendor model for procurement
class VendorBase(BaseModel):
    name: str
    mobile: Optional[str] = None
    address: Optional[str] = None
    village: Optional[str] = None
    products: List[str] = []
    is_active: bool = True

class VendorCreate(VendorBase):
    pass

class Vendor(VendorBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    total_purchases: float = 0
    total_paid: float = 0
    outstanding_dues: float = 0
    transaction_count: int = 0
    last_transaction_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

# Vendor Payment Model
class VendorPaymentCreate(BaseModel):
    vendor_id: str
    amount: float
    payment_mode: str = "cash"  # cash, online
    notes: Optional[str] = None

class VendorPayment(VendorPaymentCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Vendor Procurement Model
class VendorProcurementBase(BaseModel):
    vendor_id: str
    product_id: str
    quantity: float
    rate: float
    outlet_id: str  # Stock will be added to this outlet
    payment_mode: str = "cash"  # cash, online, credit
    cash_amount: float = 0
    online_amount: float = 0
    notes: Optional[str] = None
    # Manual entry fields for quick transactions
    manual_vendor_name: Optional[str] = None
    manual_vendor_mobile: Optional[str] = None
    manual_product_name: Optional[str] = None
    manual_product_unit: Optional[str] = None
    # Product variety (optional)
    variety_id: Optional[str] = None
    variety_name: Optional[str] = None

class VendorProcurement(VendorProcurementBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    receipt_number: str = ""
    vendor_name: Optional[str] = None
    product_name: Optional[str] = None
    outlet_name: Optional[str] = None
    total_amount: float = 0
    credit_amount: float = 0
    payment_status: str = "paid"  # paid or credit
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Notification Model
class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str  # Target user (or 'all_admins', 'all_agents', etc.)
    title: str
    message: str
    type: str  # 'request', 'approval', 'rejection', 'transfer', 'procurement', 'shareholder'
    reference_type: Optional[str] = None  # 'product_request', 'stock_transfer', 'shareholder_upgrade'
    reference_id: Optional[str] = None
    is_read: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Shareholder Upgrade Request Model
class ShareholderUpgradeRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    folio_number: Optional[str] = None  # Share certificate folio number
    share_value: Optional[float] = None  # Value of shares held
    certificate_data: Optional[str] = None  # Base64 encoded certificate image
    certificate_filename: Optional[str] = None
    status: str = "pending"  # pending, approved, rejected
    admin_remark: Optional[str] = None
    approved_by: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class FarmerPurchaseBase(BaseModel):
    farmer_id: str
    farmer_name: Optional[str] = None
    product_id: str
    product_name: Optional[str] = None
    quantity: float
    rate: float
    total_amount: float = 0
    payment_mode: str = "cash"  # cash, online, credit, partial
    payment_status: str = "paid"  # paid, credit (simplified)
    cash_amount: float = 0
    online_amount: float = 0
    credit_amount: float = 0
    notes: Optional[str] = None
    outlet_id: Optional[str] = None
    outlet_name: Optional[str] = None

class FarmerPurchaseCreate(BaseModel):
    farmer_id: str
    product_id: str
    quantity: float
    rate: float
    payment_status: str = "paid"  # paid or credit
    outlet_id: Optional[str] = None
    # Manual entry fields for quick transactions
    manual_farmer_name: Optional[str] = None
    manual_farmer_mobile: Optional[str] = None
    manual_product_name: Optional[str] = None
    manual_product_unit: Optional[str] = None

class FarmerPurchase(FarmerPurchaseBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    receipt_number: str = ""
    payment_status: str = "paid"
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CustomerPayment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    amount: float
    payment_mode: str  # cash, online
    notes: Optional[str] = None
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

class FarmerPayment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    farmer_id: str
    amount: float
    payment_mode: str  # cash, online
    notes: Optional[str] = None
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CustomerPaymentCreate(BaseModel):
    customer_id: str
    amount: float
    payment_mode: str
    notes: Optional[str] = None

class FarmerPaymentCreate(BaseModel):
    farmer_id: str
    amount: float
    payment_mode: str
    notes: Optional[str] = None

# ===================== HELPER FUNCTIONS =====================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        return {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"],
            "outlet_id": user.get("outlet_id"),
            "is_active": user.get("is_active", True)
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def generate_bill_number() -> str:
    today = datetime.utcnow()
    prefix = today.strftime("BILL%Y%m%d")
    count = await db.sales.count_documents({
        "created_at": {
            "$gte": datetime(today.year, today.month, today.day),
            "$lt": datetime(today.year, today.month, today.day) + timedelta(days=1)
        }
    })
    return f"{prefix}{str(count + 1).zfill(4)}"

async def generate_receipt_number() -> str:
    today = datetime.utcnow()
    prefix = today.strftime("RCP%Y%m%d")
    count = await db.farmer_purchases.count_documents({
        "created_at": {
            "$gte": datetime(today.year, today.month, today.day),
            "$lt": datetime(today.year, today.month, today.day) + timedelta(days=1)
        }
    })
    return f"{prefix}{str(count + 1).zfill(4)}"

# ===================== AUTHENTICATION ROUTES =====================

@api_router.post("/auth/login", response_model=Token)
async def login(request: LoginRequest):
    user = await db.users.find_one({"username": request.username})
    if not user or not verify_password(request.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is disabled")
    
    # Check if agent is pending approval
    if user.get("role") == "agent" and user.get("status") == "pending":
        raise HTTPException(status_code=401, detail="Your account is pending approval by admin")
    
    if user.get("status") == "rejected":
        raise HTTPException(status_code=401, detail="Your account registration was rejected")
    
    access_token = create_access_token(data={"sub": user["username"]})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"],
            "outlet_id": user.get("outlet_id"),
            "status": user.get("status", "active"),
            "village": user.get("village"),
            "mobile": user.get("mobile")
        }
    }

@api_router.post("/auth/register")
async def register(request: RegisterRequest):
    """Register a new user (farmer or agent)"""
    # Check if username exists
    existing = await db.users.find_one({"username": request.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Only farmer and agent can self-register
    if request.role not in ["farmer", "agent"]:
        raise HTTPException(status_code=400, detail="Invalid role. Only farmer or agent can self-register")
    
    # Determine status based on role
    status = "active" if request.role == "farmer" else "pending"  # Agents need approval
    
    user_obj = UserInDB(
        username=request.username,
        full_name=request.full_name,
        role=request.role,
        mobile=request.mobile,
        village=request.village,
        status=status,
        is_active=True if request.role == "farmer" else False,  # Agents inactive until approved
        hashed_password=get_password_hash(request.password)
    )
    
    await db.users.insert_one(user_obj.dict())
    
    if request.role == "agent":
        return {
            "message": "Registration successful. Your account is pending approval by admin.",
            "status": "pending"
        }
    else:
        # Auto-login for farmers
        access_token = create_access_token(data={"sub": request.username})
        return {
            "message": "Registration successful",
            "status": "active",
            "access_token": access_token,
            "token_type": "bearer",
            "user": {
                "id": user_obj.id,
                "username": user_obj.username,
                "full_name": user_obj.full_name,
                "role": user_obj.role,
                "outlet_id": None,
                "status": "active",
                "village": user_obj.village,
                "mobile": user_obj.mobile
            }
        }

@api_router.post("/auth/change-password")
async def change_password(request: PasswordChangeRequest, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]})
    if not verify_password(request.old_password, user["hashed_password"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    new_hash = get_password_hash(request.new_password)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"hashed_password": new_hash, "updated_at": datetime.utcnow()}}
    )
    return {"message": "Password changed successfully"}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# ===================== USER MANAGEMENT ROUTES (Admin Only) =====================

@api_router.post("/users", response_model=User)
async def create_user(user: UserCreate, current_user: dict = Depends(require_admin)):
    existing = await db.users.find_one({"username": user.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    user_dict = user.dict()
    password = user_dict.pop("password")
    user_obj = UserInDB(**user_dict, hashed_password=get_password_hash(password))
    await db.users.insert_one(user_obj.dict())
    return User(**user_dict, id=user_obj.id, created_at=user_obj.created_at, updated_at=user_obj.updated_at)

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(require_admin)):
    users = await db.users.find().to_list(1000)
    return [User(
        id=u["id"],
        username=u["username"],
        full_name=u["full_name"],
        role=u["role"],
        outlet_id=u.get("outlet_id"),
        mobile=u.get("mobile"),
        is_active=u.get("is_active", True),
        created_at=u.get("created_at", datetime.utcnow()),
        updated_at=u.get("updated_at", datetime.utcnow())
    ) for u in users]

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, updates: dict, current_user: dict = Depends(require_admin)):
    # Cannot change username
    if "username" in updates:
        del updates["username"]
    
    if "password" in updates:
        updates["hashed_password"] = get_password_hash(updates.pop("password"))
    
    updates["updated_at"] = datetime.utcnow()
    result = await db.users.update_one({"id": user_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_admin)):
    # Soft delete - just deactivate
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deactivated successfully"}

@api_router.get("/users/pending")
async def get_pending_users(current_user: dict = Depends(require_admin)):
    """Get all pending agent registrations"""
    users = await db.users.find({"status": "pending", "role": "agent"}).to_list(1000)
    return [User(
        id=u["id"],
        username=u["username"],
        full_name=u["full_name"],
        role=u["role"],
        outlet_id=u.get("outlet_id"),
        mobile=u.get("mobile"),
        village=u.get("village"),
        is_active=u.get("is_active", False),
        status=u.get("status", "pending"),
        created_at=u.get("created_at", datetime.utcnow()),
        updated_at=u.get("updated_at", datetime.utcnow())
    ) for u in users]

@api_router.post("/users/approve")
async def approve_agent(request: ApproveAgentRequest, current_user: dict = Depends(require_admin)):
    """Approve or reject agent registration and assign outlet"""
    user = await db.users.find_one({"id": request.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get("role") != "agent":
        raise HTTPException(status_code=400, detail="Only agent registrations can be approved")
    
    if request.approved:
        # Verify outlet exists
        outlet = await db.outlets.find_one({"id": request.outlet_id})
        if not outlet:
            raise HTTPException(status_code=404, detail="Outlet not found")
        
        await db.users.update_one(
            {"id": request.user_id},
            {"$set": {
                "status": "active",
                "is_active": True,
                "outlet_id": request.outlet_id,
                "updated_at": datetime.utcnow()
            }}
        )
        return {"message": "Agent approved and outlet assigned successfully"}
    else:
        await db.users.update_one(
            {"id": request.user_id},
            {"$set": {
                "status": "rejected",
                "is_active": False,
                "updated_at": datetime.utcnow()
            }}
        )
        return {"message": "Agent registration rejected"}

# ===================== PRODUCT REQUEST ROUTES (Farmer Buy/Sell Requests) =====================

@api_router.post("/product-requests")
async def create_product_request(request: ProductRequestBase, current_user: dict = Depends(get_current_user)):
    """Create a buy/sell request (for farmers)"""
    if current_user["role"] not in ["farmer", "agent"]:
        raise HTTPException(status_code=403, detail="Only farmers and agents can create product requests")
    
    # Handle custom product for sell requests
    product_name = None
    if request.product_id == 'custom' and request.custom_product_name:
        product_name = request.custom_product_name
    else:
        product = await db.products.find_one({"id": request.product_id})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        product_name = product["name"]
    
    # Get outlet name if outlet_id provided
    outlet_name = None
    if request.outlet_id:
        outlet = await db.outlets.find_one({"id": request.outlet_id})
        if outlet:
            outlet_name = outlet["name"]
    
    req_obj = ProductRequest(
        **request.dict(),
        farmer_id=current_user["id"],
        farmer_name=current_user.get("full_name"),
        product_name=product_name,
        status="pending"
    )
    
    # Add outlet_name to the document
    req_dict = req_obj.dict()
    req_dict["outlet_name"] = outlet_name
    
    await db.product_requests.insert_one(req_dict)
    return {"message": "Request submitted successfully", "request_id": req_obj.id}

@api_router.get("/product-requests")
async def get_product_requests(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get product requests - farmers see their own, agents/admins see all for their outlet"""
    query = {}
    
    if current_user["role"] == "farmer":
        query["farmer_id"] = current_user["id"]
    elif current_user["role"] == "agent":
        # Agents see requests assigned to their outlet or unassigned
        query["$or"] = [
            {"outlet_id": current_user.get("outlet_id")},
            {"outlet_id": None}
        ]
    # Admin sees all
    
    if status:
        query["status"] = status
    
    requests = await db.product_requests.find(query).sort("created_at", -1).to_list(1000)
    # Clean MongoDB _id
    result = []
    for r in requests:
        r.pop('_id', None)
        result.append(r)
    return result

@api_router.put("/product-requests/{request_id}")
async def update_product_request(request_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update a product request (approve/reject/complete) - any party can mark as resolved"""
    req = await db.product_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Allow farmer, agent, or admin to update/resolve
    if current_user["role"] == "farmer" and req.get("farmer_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["processed_by"] = current_user["id"]
    updates["processed_at"] = datetime.utcnow()
    
    if current_user["role"] == "agent" and not updates.get("outlet_id"):
        updates["outlet_id"] = current_user.get("outlet_id")
    
    await db.product_requests.update_one({"id": request_id}, {"$set": updates})
    return {"message": "Request updated successfully"}

@api_router.get("/my-transactions")
async def get_my_transactions(current_user: dict = Depends(get_current_user)):
    """Get transactions for current user (farmers see their sales/purchases)"""
    
    def clean_docs(docs):
        """Remove MongoDB _id from documents"""
        result = []
        for d in docs:
            if d:
                d.pop('_id', None)
                result.append(d)
        return result
    
    if current_user["role"] == "farmer":
        # Get farmer purchases (when FPO bought from this farmer)
        farmer = await db.farmers.find_one({"mobile": current_user.get("mobile")})
        purchases = []
        if farmer:
            farmer.pop('_id', None)
            purchases_raw = await db.farmer_purchases.find({"farmer_id": farmer["id"]}).sort("created_at", -1).to_list(100)
            purchases = clean_docs(purchases_raw)
        
        # Get product requests
        requests_raw = await db.product_requests.find({"farmer_id": current_user["id"]}).sort("created_at", -1).to_list(100)
        requests = clean_docs(requests_raw)
        
        return {
            "purchases": purchases,
            "requests": requests,
            "farmer_profile": farmer
        }
    elif current_user["role"] == "agent":
        # Agents see their outlet transactions plus their own farmer transactions
        outlet_id = current_user.get("outlet_id")
        
        sales_raw = await db.sales.find({"outlet_id": outlet_id}).sort("created_at", -1).to_list(100) if outlet_id else []
        sales = clean_docs(sales_raw)
        
        # Also check if agent is registered as farmer
        farmer = await db.farmers.find_one({"mobile": current_user.get("mobile")})
        farmer_purchases = []
        if farmer:
            farmer.pop('_id', None)
            purchases_raw = await db.farmer_purchases.find({"farmer_id": farmer["id"]}).sort("created_at", -1).to_list(100)
            farmer_purchases = clean_docs(purchases_raw)
        
        return {
            "outlet_sales": sales,
            "farmer_purchases": farmer_purchases,
            "farmer_profile": farmer
        }
    
    return {"purchases": [], "requests": [], "message": "No transactions"}

# ===================== OUTLET ROUTES =====================

@api_router.post("/outlets", response_model=Outlet)
async def create_outlet(outlet: OutletCreate, current_user: dict = Depends(require_admin)):
    outlet_obj = Outlet(**outlet.dict())
    await db.outlets.insert_one(outlet_obj.dict())
    return outlet_obj

@api_router.get("/outlets", response_model=List[Outlet])
async def get_outlets(current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["id"] = current_user["outlet_id"]
    outlets = await db.outlets.find(query).to_list(1000)
    return [Outlet(**o) for o in outlets]

# IMPORTANT: These specific routes must come BEFORE /outlets/{outlet_id} 
# to avoid path parameter conflict

@api_router.get("/outlets/ledger")
async def get_all_outlets_ledger(current_user: dict = Depends(require_admin)):
    """Get ledger summary for all outlets (Admin only)"""
    outlets = await db.outlets.find({"is_active": True}).to_list(100)
    
    ledger_summary = []
    for outlet in outlets:
        outlet_id = outlet["id"]
        
        # Get total sales at this outlet (cash + online collected)
        sales = await db.sales.find({
            "outlet_id": outlet_id,
            "$or": [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]
        }).to_list(10000)
        
        total_sales = sum(s.get("total_amount", 0) for s in sales)
        total_cash_collected = sum(s.get("cash_amount", 0) for s in sales)
        total_online_collected = sum(s.get("online_amount", 0) for s in sales)
        total_credit_given = sum(s.get("credit_amount", 0) for s in sales)
        
        # Get total deposits made by this outlet
        deposits = await db.outlet_cash_deposits.find({
            "outlet_id": outlet_id,
            "status": "confirmed"
        }).to_list(10000)
        
        total_deposited = sum(d.get("amount", 0) for d in deposits)
        
        # Calculate outstanding (cash collected but not deposited)
        cash_outstanding = total_cash_collected - total_deposited
        
        # Get pending deposits
        pending_deposits = await db.outlet_cash_deposits.find({
            "outlet_id": outlet_id,
            "status": "pending"
        }).to_list(100)
        
        pending_deposit_amount = sum(d.get("amount", 0) for d in pending_deposits)
        
        ledger_summary.append({
            "outlet_id": outlet_id,
            "outlet_name": outlet["name"],
            "outlet_address": outlet.get("address", ""),
            "total_sales": total_sales,
            "total_cash_collected": total_cash_collected,
            "total_online_collected": total_online_collected,
            "total_credit_given": total_credit_given,
            "total_deposited": total_deposited,
            "cash_outstanding": cash_outstanding,
            "pending_deposit_amount": pending_deposit_amount,
            "transaction_count": len(sales)
        })
    
    # Calculate totals
    totals = {
        "total_sales": sum(o["total_sales"] for o in ledger_summary),
        "total_cash_collected": sum(o["total_cash_collected"] for o in ledger_summary),
        "total_online_collected": sum(o["total_online_collected"] for o in ledger_summary),
        "total_credit_given": sum(o["total_credit_given"] for o in ledger_summary),
        "total_deposited": sum(o["total_deposited"] for o in ledger_summary),
        "cash_outstanding": sum(o["cash_outstanding"] for o in ledger_summary),
        "pending_deposit_amount": sum(o["pending_deposit_amount"] for o in ledger_summary)
    }
    
    return {
        "outlets": ledger_summary,
        "totals": totals
    }

@api_router.post("/outlets/cash-deposit")
async def create_cash_deposit(
    deposit: OutletCashDepositCreate,
    current_user: dict = Depends(get_current_user)
):
    """Record a cash deposit from outlet to admin"""
    outlet = await db.outlets.find_one({"id": deposit.outlet_id})
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    
    # Get user name
    user = await db.users.find_one({"id": current_user["id"]})
    user_name = user.get("full_name", user.get("username", "Unknown")) if user else "Unknown"
    
    deposit_obj = OutletCashDeposit(
        outlet_id=deposit.outlet_id,
        outlet_name=outlet["name"],
        amount=deposit.amount,
        payment_mode=deposit.payment_mode,
        reference_number=deposit.reference_number,
        notes=deposit.notes,
        deposited_by=current_user["id"],
        deposited_by_name=user_name,
        status="pending" if current_user["role"] != "admin" else "confirmed",
        received_by=current_user["id"] if current_user["role"] == "admin" else None,
        received_by_name=user_name if current_user["role"] == "admin" else None,
        confirmed_at=datetime.utcnow() if current_user["role"] == "admin" else None
    )
    
    await db.outlet_cash_deposits.insert_one(deposit_obj.dict())
    
    return {
        "message": "Deposit recorded successfully",
        "status": deposit_obj.status,
        "deposit_id": deposit_obj.id
    }

@api_router.get("/outlets/cash-deposits")
async def get_cash_deposits(
    outlet_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get cash deposit records"""
    query = {}
    
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    if status:
        query["status"] = status
    
    deposits = await db.outlet_cash_deposits.find(query).sort("created_at", -1).to_list(500)
    
    # Clean MongoDB _id
    for d in deposits:
        d.pop('_id', None)
    
    return deposits

@api_router.get("/outlets/{outlet_id}", response_model=Outlet)
async def get_outlet(outlet_id: str, current_user: dict = Depends(get_current_user)):
    outlet = await db.outlets.find_one({"id": outlet_id})
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    return Outlet(**outlet)

@api_router.put("/outlets/{outlet_id}")
async def update_outlet(outlet_id: str, updates: dict, current_user: dict = Depends(require_admin)):
    updates["updated_at"] = datetime.utcnow()
    result = await db.outlets.update_one({"id": outlet_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Outlet not found")
    return {"message": "Outlet updated successfully"}

@api_router.delete("/outlets/{outlet_id}")
async def delete_outlet(outlet_id: str, current_user: dict = Depends(require_admin)):
    result = await db.outlets.update_one(
        {"id": outlet_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Outlet not found")
    return {"message": "Outlet deactivated successfully"}

# ===================== PRODUCT ROUTES =====================

@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate, current_user: dict = Depends(require_admin)):
    product_obj = Product(**product.dict())
    await db.products.insert_one(product_obj.dict())
    return product_obj

@api_router.get("/products", response_model=List[Product])
async def get_products(category: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    if category:
        query["category"] = category
    products = await db.products.find(query).to_list(1000)
    return [Product(**p) for p in products]

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return Product(**product)

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, updates: dict, current_user: dict = Depends(require_admin)):
    updates["updated_at"] = datetime.utcnow()
    result = await db.products.update_one({"id": product_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product updated successfully"}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(require_admin)):
    result = await db.products.update_one(
        {"id": product_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deactivated successfully"}

# ===================== STOCK ROUTES =====================

@api_router.get("/stock")
async def get_stock(outlet_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    stocks = await db.stock.find(query).to_list(1000)
    
    # Batch fetch products and outlets to avoid N+1 query problem
    product_ids = list(set(s["product_id"] for s in stocks))
    outlet_ids = list(set(s["outlet_id"] for s in stocks))
    
    products_list = await db.products.find({"id": {"$in": product_ids}}).to_list(1000)
    outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(1000)
    
    products_map = {p["id"]: p for p in products_list}
    outlets_map = {o["id"]: o for o in outlets_list}
    
    # Enrich with product and outlet info using cached data
    result = []
    for s in stocks:
        s.pop('_id', None)  # Remove MongoDB _id
        product = products_map.get(s["product_id"])
        outlet = outlets_map.get(s["outlet_id"])
        result.append({
            **s,
            "product_name": product["name"] if product else "Unknown",
            "product_unit": product["unit"] if product else "",
            "outlet_name": outlet["name"] if outlet else "Unknown"
        })
    return result

@api_router.get("/stock/consolidated")
async def get_consolidated_stock(current_user: dict = Depends(require_admin)):
    """Get consolidated stock across all outlets for bulk orders"""
    pipeline = [
        {"$group": {
            "_id": "$product_id",
            "total_quantity": {"$sum": "$quantity"},
            "total_opening": {"$sum": "$opening_stock"},
            "total_received": {"$sum": "$stock_received"},
            "total_sold": {"$sum": "$stock_sold"},
            "total_damaged": {"$sum": "$stock_damaged"}
        }}
    ]
    results = await db.stock.aggregate(pipeline).to_list(1000)
    
    # Batch fetch products to avoid N+1 query problem
    product_ids = [r["_id"] for r in results]
    products_list = await db.products.find({"id": {"$in": product_ids}}).to_list(1000)
    products_map = {p["id"]: p for p in products_list}
    
    # Enrich with product info using cached data
    consolidated = []
    for r in results:
        product = products_map.get(r["_id"])
        if product:
            consolidated.append({
                "product_id": r["_id"],
                "product_name": product["name"],
                "product_unit": product["unit"],
                "total_quantity": r["total_quantity"],
                "total_opening": r["total_opening"],
                "total_received": r["total_received"],
                "total_sold": r["total_sold"],
                "total_damaged": r["total_damaged"]
            })
    return consolidated

@api_router.post("/stock/add")
async def add_stock(request: StockAddRequest, current_user: dict = Depends(get_current_user)):
    """Add stock to an outlet"""
    # Check if stock record exists
    existing = await db.stock.find_one({
        "product_id": request.product_id,
        "outlet_id": request.outlet_id
    })
    
    if existing:
        await db.stock.update_one(
            {"id": existing["id"]},
            {"$inc": {
                "quantity": request.quantity,
                "stock_received": request.quantity
            }, "$set": {"updated_at": datetime.utcnow()}}
        )
    else:
        stock = Stock(
            product_id=request.product_id,
            outlet_id=request.outlet_id,
            quantity=request.quantity,
            opening_stock=request.quantity,
            stock_received=request.quantity
        )
        await db.stock.insert_one(stock.dict())
    
    # Log movement
    movement = StockMovement(
        product_id=request.product_id,
        to_outlet_id=request.outlet_id,
        quantity=request.quantity,
        movement_type="addition",
        created_by=current_user["id"]
    )
    await db.stock_movements.insert_one(movement.dict())
    
    return {"message": "Stock added successfully"}

@api_router.post("/stock/allocate")
async def allocate_stock(request: StockAllocationRequest, current_user: dict = Depends(require_admin)):
    """Allocate stock from one outlet to another"""
    # Check source stock
    source = await db.stock.find_one({
        "product_id": request.product_id,
        "outlet_id": request.from_outlet_id
    })
    
    if not source or source["quantity"] < request.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock at source outlet")
    
    # Deduct from source
    await db.stock.update_one(
        {"id": source["id"]},
        {"$inc": {"quantity": -request.quantity}, "$set": {"updated_at": datetime.utcnow()}}
    )
    
    # Add to destination
    dest = await db.stock.find_one({
        "product_id": request.product_id,
        "outlet_id": request.to_outlet_id
    })
    
    if dest:
        await db.stock.update_one(
            {"id": dest["id"]},
            {"$inc": {
                "quantity": request.quantity,
                "stock_received": request.quantity
            }, "$set": {"updated_at": datetime.utcnow()}}
        )
    else:
        stock = Stock(
            product_id=request.product_id,
            outlet_id=request.to_outlet_id,
            quantity=request.quantity,
            stock_received=request.quantity
        )
        await db.stock.insert_one(stock.dict())
    
    # Log movement
    movement = StockMovement(
        product_id=request.product_id,
        from_outlet_id=request.from_outlet_id,
        to_outlet_id=request.to_outlet_id,
        quantity=request.quantity,
        movement_type="allocation",
        created_by=current_user["id"]
    )
    await db.stock_movements.insert_one(movement.dict())
    
    return {"message": "Stock allocated successfully"}

@api_router.post("/stock/damage")
async def report_damage(request: StockDamageRequest, current_user: dict = Depends(get_current_user)):
    """Report damaged/discarded stock with mandatory reason"""
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="A valid reason is required for damaged stock")
    
    # Check stock
    stock = await db.stock.find_one({
        "product_id": request.product_id,
        "outlet_id": request.outlet_id
    })
    
    if not stock or stock["quantity"] < request.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock")
    
    # Update stock
    await db.stock.update_one(
        {"id": stock["id"]},
        {"$inc": {
            "quantity": -request.quantity,
            "stock_damaged": request.quantity
        }, "$set": {"updated_at": datetime.utcnow()}}
    )
    
    # Log movement
    movement = StockMovement(
        product_id=request.product_id,
        from_outlet_id=request.outlet_id,
        quantity=request.quantity,
        movement_type="damage",
        reason=request.reason,
        created_by=current_user["id"]
    )
    await db.stock_movements.insert_one(movement.dict())
    
    return {"message": "Damage reported successfully"}

# ===================== STOCK TRANSFER REQUEST SYSTEM =====================

@api_router.post("/stock/transfer-request")
async def create_transfer_request(request: StockTransferRequestCreate, current_user: dict = Depends(get_current_user)):
    """Create a stock transfer request (any user can request, admin approval needed)"""
    # Get product details
    product = await db.products.find_one({"id": request.product_id, "is_active": True})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get outlet details
    from_outlet = await db.outlets.find_one({"id": request.from_outlet_id, "is_active": True})
    to_outlet = await db.outlets.find_one({"id": request.to_outlet_id, "is_active": True})
    
    if not from_outlet or not to_outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    
    if request.from_outlet_id == request.to_outlet_id:
        raise HTTPException(status_code=400, detail="Source and destination outlets must be different")
    
    # Check available stock
    stock = await db.stock.find_one({
        "product_id": request.product_id,
        "outlet_id": request.from_outlet_id
    })
    
    if not stock or stock["quantity"] < request.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock in source outlet")
    
    # Get user details
    user = await db.users.find_one({"id": current_user["id"]})
    
    # Create transfer request
    transfer_request = StockTransferRequest(
        product_id=request.product_id,
        product_name=product["name"],
        from_outlet_id=request.from_outlet_id,
        from_outlet_name=from_outlet["name"],
        to_outlet_id=request.to_outlet_id,
        to_outlet_name=to_outlet["name"],
        quantity=request.quantity,
        reason=request.reason,
        status="pending",
        requested_by=current_user["id"],
        requested_by_name=user.get("full_name", user.get("username", "Unknown"))
    )
    
    await db.stock_transfer_requests.insert_one(transfer_request.dict())
    
    return {"message": "Transfer request created successfully", "request_id": transfer_request.id}

@api_router.get("/stock/transfer-requests")
async def get_transfer_requests(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get stock transfer requests - Admin sees all, others see their own"""
    query = {}
    
    if current_user["role"] != "admin":
        query["requested_by"] = current_user["id"]
    
    if status:
        query["status"] = status
    
    requests = await db.stock_transfer_requests.find(query).sort("created_at", -1).to_list(500)
    
    # Remove _id from results
    result = []
    for r in requests:
        r.pop('_id', None)
        result.append(r)
    
    return result

@api_router.get("/stock/transfer-requests/pending-count")
async def get_pending_transfer_count(current_user: dict = Depends(require_admin)):
    """Get count of pending transfer requests - Admin only"""
    count = await db.stock_transfer_requests.count_documents({"status": "pending"})
    return {"count": count}

@api_router.put("/stock/transfer-requests/{request_id}/approve")
async def approve_transfer_request(
    request_id: str, 
    remark: Optional[str] = None, 
    approved_quantity: Optional[float] = None,
    current_user: dict = Depends(require_admin)
):
    """Approve a stock transfer request - Admin only. Can approve partial or modified quantity."""
    # Find the request
    transfer = await db.stock_transfer_requests.find_one({"id": request_id})
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer request not found")
    
    if transfer["status"] != "pending":
        raise HTTPException(status_code=400, detail="This request has already been processed")
    
    # Use approved_quantity if provided, otherwise use original quantity
    final_quantity = approved_quantity if approved_quantity is not None else transfer["quantity"]
    
    if final_quantity <= 0:
        raise HTTPException(status_code=400, detail="Approved quantity must be greater than 0")
    
    # Check stock again before approving
    stock = await db.stock.find_one({
        "product_id": transfer["product_id"],
        "outlet_id": transfer["from_outlet_id"]
    })
    
    if not stock or stock["quantity"] < final_quantity:
        raise HTTPException(status_code=400, detail=f"Insufficient stock. Available: {stock['quantity'] if stock else 0}")
    
    # Get admin details
    admin = await db.users.find_one({"id": current_user["id"]})
    
    # Perform the actual stock transfer with final_quantity
    # Deduct from source
    await db.stock.update_one(
        {"id": stock["id"]},
        {"$inc": {"quantity": -final_quantity}, "$set": {"updated_at": datetime.utcnow()}}
    )
    
    # Add to destination
    dest_stock = await db.stock.find_one({
        "product_id": transfer["product_id"],
        "outlet_id": transfer["to_outlet_id"]
    })
    
    if dest_stock:
        await db.stock.update_one(
            {"id": dest_stock["id"]},
            {"$inc": {
                "quantity": final_quantity,
                "stock_received": final_quantity
            }, "$set": {"updated_at": datetime.utcnow()}}
        )
    else:
        new_stock = Stock(
            product_id=transfer["product_id"],
            outlet_id=transfer["to_outlet_id"],
            quantity=final_quantity,
            stock_received=final_quantity
        )
        await db.stock.insert_one(new_stock.dict())
    
    # Determine approval type for remark
    approval_note = ""
    if approved_quantity is not None and approved_quantity != transfer["quantity"]:
        if approved_quantity < transfer["quantity"]:
            approval_note = f"Partial approval: {final_quantity} of {transfer['quantity']} requested. "
        else:
            approval_note = f"Increased allocation: {final_quantity} (requested {transfer['quantity']}). "
    
    # Log the movement
    movement = StockMovement(
        product_id=transfer["product_id"],
        from_outlet_id=transfer["from_outlet_id"],
        to_outlet_id=transfer["to_outlet_id"],
        quantity=final_quantity,
        movement_type="transfer_approved",
        reason=f"{approval_note}Transfer request approved. {remark or ''}".strip(),
        created_by=current_user["id"]
    )
    await db.stock_movements.insert_one(movement.dict())
    
    # Update the request status
    await db.stock_transfer_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_quantity": final_quantity,
            "approved_by": current_user["id"],
            "approved_by_name": admin.get("full_name", admin.get("username", "Admin")),
            "admin_remark": f"{approval_note}{remark or ''}".strip() or "Approved",
            "updated_at": datetime.utcnow()
        }}
    )
    
    # Notify the requester
    await create_notification(
        transfer["requested_by"],
        "Stock Transfer Approved!",
        f"Your request for {transfer['product_name']} has been approved. {approval_note}Transferred: {final_quantity} units.",
        "approval",
        "stock_transfer",
        request_id
    )
    
    return {
        "message": "Transfer request approved and stock transferred successfully",
        "approved_quantity": final_quantity,
        "requested_quantity": transfer["quantity"]
    }

@api_router.put("/stock/transfer-requests/{request_id}/reject")
async def reject_transfer_request(request_id: str, remark: str = "", current_user: dict = Depends(require_admin)):
    """Reject a stock transfer request - Admin only"""
    # Find the request
    transfer = await db.stock_transfer_requests.find_one({"id": request_id})
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer request not found")
    
    if transfer["status"] != "pending":
        raise HTTPException(status_code=400, detail="This request has already been processed")
    
    # Get admin details
    admin = await db.users.find_one({"id": current_user["id"]})
    
    # Update the request status
    await db.stock_transfer_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["id"],
            "approved_by_name": admin.get("full_name", admin.get("username", "Admin")),
            "admin_remark": remark or "Request rejected",
            "updated_at": datetime.utcnow()
        }}
    )
    
    return {"message": "Transfer request rejected"}

# ===================== CUSTOMER ROUTES (KHATA/LEDGER SYSTEM) =====================

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    """Create a new customer with optional shareholder/farmer linking"""
    customer_obj = Customer(**customer.dict())
    await db.customers.insert_one(customer_obj.dict())
    return customer_obj

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(
    search: Optional[str] = None,
    customer_type: Optional[str] = None,
    has_dues: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get customers with search and filter support"""
    query = {"is_active": True}
    
    # Search by name, mobile, or village (case-insensitive partial match)
    if search:
        import re
        search_regex = re.compile(search, re.IGNORECASE)
        query["$or"] = [
            {"name": {"$regex": search_regex}},
            {"mobile": {"$regex": search_regex}},
            {"village": {"$regex": search_regex}}
        ]
    
    # Filter by customer type
    if customer_type:
        query["customer_type"] = customer_type
    
    # Filter customers with outstanding dues
    if has_dues:
        query["outstanding_balance"] = {"$gt": 0}
    
    customers = await db.customers.find(query).sort("name", 1).to_list(1000)
    return [Customer(**c) for c in customers]

@api_router.get("/customers/search")
async def search_customers(
    q: str,
    current_user: dict = Depends(get_current_user)
):
    """Quick search endpoint for customers - used in autocomplete"""
    if not q or len(q.strip()) < 1:
        return []

    import re
    # Escape special regex chars so user input like "(", "+" doesn't crash the query
    safe_q = re.escape(q.strip())
    search_regex = {"$regex": safe_q, "$options": "i"}
    query = {
        "is_active": True,
        "$or": [
            {"name": search_regex},
            {"name_hi": search_regex},
            {"mobile": search_regex},
            {"village": search_regex},
            {"address": search_regex},
            {"folio_number": search_regex},
        ]
    }

    customers = await db.customers.find(query).limit(20).to_list(20)
    return [{
        "id": c["id"],
        "name": c["name"],
        "name_hi": c.get("name_hi", ""),
        "mobile": c.get("mobile", ""),
        "village": c.get("village", ""),
        "address": c.get("address", ""),
        "customer_type": c.get("customer_type", "walk_in"),
        "folio_number": c.get("folio_number", ""),
        "outstanding_balance": c.get("outstanding_balance", 0)
    } for c in customers]

@api_router.get("/customers/with-dues")
async def get_customers_with_dues(current_user: dict = Depends(get_current_user)):
    """Get all customers with outstanding dues (for Khata overview)"""
    customers = await db.customers.find({
        "is_active": True,
        "outstanding_balance": {"$gt": 0}
    }).sort("outstanding_balance", -1).to_list(1000)
    
    total_dues = sum(c.get("outstanding_balance", 0) for c in customers)
    
    return {
        "customers": [Customer(**c) for c in customers],
        "total_dues": total_dues,
        "count": len(customers)
    }

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return Customer(**customer)

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update customer - including upgrading to shareholder"""
    updates["updated_at"] = datetime.utcnow()
    result = await db.customers.update_one({"id": customer_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer updated successfully"}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(require_admin)):
    """Soft delete a customer"""
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

@api_router.get("/customers/{customer_id}/ledger")
async def get_customer_ledger(
    customer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive customer ledger (Khata) with all transactions"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Build date filter
    date_filter = {}
    if start_date:
        date_filter["$gte"] = datetime.fromisoformat(start_date)
    if end_date:
        date_filter["$lte"] = datetime.fromisoformat(end_date)
    
    # Get sales for this customer
    sales_query = {"customer_id": customer_id}
    if date_filter:
        sales_query["created_at"] = date_filter
    
    sales_raw = await db.sales.find(sales_query).sort("created_at", -1).to_list(1000)
    
    # Get payments for this customer
    payments_query = {"customer_id": customer_id}
    if date_filter:
        payments_query["created_at"] = date_filter
    
    payments_raw = await db.customer_payments.find(payments_query).sort("created_at", -1).to_list(1000)
    
    # Clean MongoDB _id from results
    def clean_doc(doc):
        doc.pop('_id', None)
        return doc
    
    sales = [clean_doc(s) for s in sales_raw]
    payments = [clean_doc(p) for p in payments_raw]

    # Split active vs cancelled sales — only active sales count in totals
    active_sales = [s for s in sales if not s.get("is_deleted", False)]
    cancelled_sales = [s for s in sales if s.get("is_deleted", False)]

    # Calculate ledger summary (cancelled sales are excluded from totals)
    total_billed = sum(s.get("total_amount", 0) for s in active_sales)
    total_credit_given = sum(s.get("credit_amount", 0) for s in active_sales)
    total_payments = sum(p.get("amount", 0) for p in payments)

    # Build transaction timeline (combined sales and payments sorted by date)
    transactions = []
    for s in sales:
        is_cancelled = s.get("is_deleted", False)
        transactions.append({
            "id": s["id"],
            "type": "sale",
            "date": s["created_at"],
            "reference": s.get("bill_number", ""),
            "description": f"Bill #{s.get('bill_number', 'N/A')}" + (" (CANCELLED)" if is_cancelled else ""),
            "debit": 0 if is_cancelled else s.get("credit_amount", 0),
            "credit": 0,
            "items_count": len(s.get("items", [])),
            "total_amount": s.get("total_amount", 0),
            "payment_mode": s.get("payment_mode", ""),
            "is_cancelled": is_cancelled,
            "deleted_at": s.get("deleted_at"),
            "deletion_reason": s.get("deletion_reason", ""),
        })
    
    for p in payments:
        transactions.append({
            "id": p["id"],
            "type": "payment",
            "date": p["created_at"],
            "reference": "",
            "description": f"Payment - {p.get('payment_mode', 'cash').upper()}",
            "debit": 0,
            "credit": p.get("amount", 0),  # Payment reduces dues
            "payment_mode": p.get("payment_mode", "cash"),
            "notes": p.get("notes", "")
        })
    
    # Sort by date descending
    transactions.sort(key=lambda x: x["date"], reverse=True)
    
    # Remove _id from customer
    customer.pop('_id', None)
    
    return {
        "customer": Customer(**customer),
        "transactions": transactions,
        "summary": {
            "total_transactions": len(sales),
            "total_billed": total_billed,
            "total_credit_given": total_credit_given,
            "total_payments": total_payments,
            "outstanding_balance": customer.get("outstanding_balance", 0)
        },
        "sales": sales,
        "payments": payments
    }

@api_router.post("/customers/payment")
async def record_customer_payment(payment: CustomerPaymentCreate, current_user: dict = Depends(get_current_user)):
    """Record a payment from customer - adjusts dues automatically"""
    customer = await db.customers.find_one({"id": payment.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    payment_obj = CustomerPayment(
        customer_id=payment.customer_id,
        amount=payment.amount,
        payment_mode=payment.payment_mode,
        notes=payment.notes,
        created_by=current_user["id"]
    )
    await db.customer_payments.insert_one(payment_obj.dict())
    
    # Update customer balance - auto-adjust dues
    new_balance = max(0, customer.get("outstanding_balance", 0) - payment.amount)
    await db.customers.update_one(
        {"id": payment.customer_id},
        {
            "$inc": {"total_paid": payment.amount},
            "$set": {
                "outstanding_balance": new_balance,
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    return {
        "message": "Payment recorded successfully",
        "new_balance": new_balance
    }

@api_router.put("/customers/{customer_id}/upgrade-shareholder")
async def upgrade_customer_to_shareholder(
    customer_id: str,
    folio_number: str,
    current_user: dict = Depends(require_admin)
):
    """Upgrade a customer to shareholder status"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {
            "customer_type": "shareholder",
            "folio_number": folio_number,
            "updated_at": datetime.utcnow()
        }}
    )
    
    return {"message": "Customer upgraded to shareholder successfully"}

# ===================== SALES ROUTES =====================

@api_router.post("/sales", response_model=Sale)
async def create_sale(sale: SaleCreate, current_user: dict = Depends(get_current_user)):
    """Create a new sale and generate bill"""
    # Verify outlet access
    if current_user["role"] == "agent" and current_user.get("outlet_id") != sale.outlet_id:
        raise HTTPException(status_code=403, detail="You can only make sales at your assigned outlet")
    
    # Check stock availability for all items
    for item in sale.items:
        stock = await db.stock.find_one({
            "product_id": item.product_id,
            "outlet_id": sale.outlet_id
        })
        if not stock or stock["quantity"] < item.quantity:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient stock for {item.product_name}"
            )
    
    # Generate bill number
    bill_number = await generate_bill_number()
    
    # Create sale
    sale_obj = Sale(
        **sale.dict(),
        bill_number=bill_number,
        created_by=current_user["id"]
    )
    await db.sales.insert_one(sale_obj.dict())
    
    # Deduct stock for each item
    for item in sale.items:
        await db.stock.update_one(
            {"product_id": item.product_id, "outlet_id": sale.outlet_id},
            {"$inc": {
                "quantity": -item.quantity,
                "stock_sold": item.quantity
            }, "$set": {"updated_at": datetime.utcnow()}}
        )
        
        # Log movement
        movement = StockMovement(
            product_id=item.product_id,
            from_outlet_id=sale.outlet_id,
            quantity=item.quantity,
            movement_type="sale",
            created_by=current_user["id"]
        )
        await db.stock_movements.insert_one(movement.dict())
    
    # Update customer ledger if credit involved
    if sale.customer_id and sale.credit_amount > 0:
        await db.customers.update_one(
            {"id": sale.customer_id},
            {"$inc": {
                "total_purchases": sale.total_amount,
                "total_credit": sale.credit_amount,
                "total_paid": sale.cash_amount + sale.online_amount,
                "outstanding_balance": sale.credit_amount,
                "transaction_count": 1
            }, "$set": {
                "updated_at": datetime.utcnow(),
                "last_transaction_date": datetime.utcnow()
            }}
        )
    elif sale.customer_id:
        await db.customers.update_one(
            {"id": sale.customer_id},
            {"$inc": {
                "total_purchases": sale.total_amount,
                "total_paid": sale.total_amount,
                "transaction_count": 1
            }, "$set": {
                "updated_at": datetime.utcnow(),
                "last_transaction_date": datetime.utcnow()
            }}
        )
    
    return sale_obj

@api_router.get("/sales")
async def get_sales(
    outlet_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_deleted: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get sales with optional filters"""
    query = {}
    
    # By default, exclude deleted transactions
    if not include_deleted:
        query["$or"] = [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]
    
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}
    
    sales = await db.sales.find(query).sort("created_at", -1).to_list(1000)
    
    # Batch fetch outlets to avoid N+1 query
    outlet_ids = list(set(s.get("outlet_id") for s in sales if s.get("outlet_id")))
    outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(1000) if outlet_ids else []
    outlets_map = {o["id"]: o for o in outlets_list}
    
    # Clean up and enrich with outlet info
    result = []
    for sale in sales:
        sale.pop('_id', None)  # Remove MongoDB _id
        outlet = outlets_map.get(sale.get("outlet_id"))
        sale["outlet_name"] = outlet["name"] if outlet else "Unknown"
        result.append(sale)
    
    return result

@api_router.get("/sales/{sale_id}")
async def get_sale(sale_id: str, current_user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    sale.pop('_id', None)
    outlet = await db.outlets.find_one({"id": sale["outlet_id"]})
    sale["outlet_name"] = outlet["name"] if outlet else "Unknown"
    sale["outlet_address"] = outlet["address"] if outlet else ""

    return sale

# ===================== TRANSACTION DELETION WITH AUTO-REVERSAL =====================

@api_router.delete("/sales/{sale_id}")
async def delete_sale(
    sale_id: str, 
    reason: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    """
    Delete a sales transaction with automatic reversal of:
    1. Inventory/Stock - Add sold products back to stock
    2. Customer Ledger - Adjust outstanding dues and totals
    
    Constraints:
    - Only admin can delete
    - Only transactions within 30 days can be deleted
    - Creates audit log for tracking
    """
    # Find the sale
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # Check if already deleted
    if sale.get("is_deleted", False):
        raise HTTPException(status_code=400, detail="Transaction already deleted")
    
    # Check 30-day constraint
    sale_date = sale.get("created_at")
    if isinstance(sale_date, str):
        sale_date = datetime.fromisoformat(sale_date)
    
    days_old = (datetime.utcnow() - sale_date).days
    if days_old > 30:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete transaction older than 30 days. This transaction is {days_old} days old."
        )
    
    reversal_details = {
        "stock_restored": [],
        "customer_ledger_adjusted": False,
        "customer_id": None,
        "credit_reversed": 0,
        "paid_amount_reversed": 0
    }
    
    # 1. REVERSE INVENTORY (Add products back to stock)
    outlet_id = sale.get("outlet_id")
    items = sale.get("items", [])
    
    for item in items:
        product_id = item.get("product_id")
        quantity = item.get("quantity", 0)
        
        if product_id and quantity > 0:
            # Find and update stock
            stock = await db.stock.find_one({
                "outlet_id": outlet_id,
                "product_id": product_id
            })
            
            if stock:
                # Add quantity back to stock
                await db.stock.update_one(
                    {"id": stock["id"]},
                    {
                        "$inc": {"quantity": quantity, "sold_quantity": -quantity},
                        "$set": {"updated_at": datetime.utcnow()}
                    }
                )
                reversal_details["stock_restored"].append({
                    "product_id": product_id,
                    "product_name": item.get("product_name", "Unknown"),
                    "quantity_restored": quantity
                })
    
    # 2. ADJUST CUSTOMER LEDGER (if credit was involved)
    customer_id = sale.get("customer_id")
    credit_amount = sale.get("credit_amount", 0)
    total_amount = sale.get("total_amount", 0)
    cash_amount = sale.get("cash_amount", 0)
    online_amount = sale.get("online_amount", 0)
    
    if customer_id:
        customer = await db.customers.find_one({"id": customer_id})
        if customer:
            # Reverse the ledger entries
            update_ops = {
                "$inc": {
                    "total_purchases": -total_amount,
                    "transaction_count": -1
                },
                "$set": {"updated_at": datetime.utcnow()}
            }
            
            # Reverse credit if any
            if credit_amount > 0:
                update_ops["$inc"]["total_credit"] = -credit_amount
                update_ops["$inc"]["outstanding_balance"] = -credit_amount
                reversal_details["credit_reversed"] = credit_amount
            
            # Reverse paid amount
            paid_amount = cash_amount + online_amount
            if paid_amount > 0:
                update_ops["$inc"]["total_paid"] = -paid_amount
                reversal_details["paid_amount_reversed"] = paid_amount
            
            await db.customers.update_one({"id": customer_id}, update_ops)
            reversal_details["customer_ledger_adjusted"] = True
            reversal_details["customer_id"] = customer_id
    
    # 3. MARK SALE AS DELETED (soft delete - preserve for audit)
    await db.sales.update_one(
        {"id": sale_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.utcnow(),
            "deleted_by": current_user["id"],
            "deletion_reason": reason
        }}
    )
    
    # 4. CREATE AUDIT LOG
    # Get admin name
    admin = await db.users.find_one({"id": current_user["id"]})
    admin_name = admin.get("full_name", admin.get("username", "Admin")) if admin else "Admin"
    
    # Clean sale data for storage
    sale.pop('_id', None)
    
    audit_log = TransactionDeletionLog(
        transaction_type="sale",
        transaction_id=sale_id,
        original_data=sale,
        bill_number=sale.get("bill_number"),
        customer_id=customer_id,
        customer_name=sale.get("customer_name"),
        total_amount=total_amount,
        deleted_by=current_user["id"],
        deleted_by_name=admin_name,
        deletion_reason=reason,
        reversal_details=reversal_details
    )
    
    await db.transaction_deletion_logs.insert_one(audit_log.dict())
    
    return {
        "message": "Transaction deleted successfully",
        "sale_id": sale_id,
        "bill_number": sale.get("bill_number"),
        "reversal_details": reversal_details
    }

@api_router.get("/deleted-transactions")
async def get_deleted_transactions(
    transaction_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    """Get audit log of all deleted transactions (Admin only)"""
    query = {}
    
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}
    
    logs = await db.transaction_deletion_logs.find(query).sort("created_at", -1).to_list(500)
    
    # Clean MongoDB _id
    for log in logs:
        log.pop('_id', None)
    
    return logs

@api_router.get("/sales/{sale_id}/can-delete")
async def check_sale_deletable(sale_id: str, current_user: dict = Depends(get_current_user)):
    """Check if a sale can be deleted (within 30 days and not already deleted)"""
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # Check if already deleted
    if sale.get("is_deleted", False):
        return {
            "can_delete": False,
            "reason": "Transaction already deleted",
            "is_admin_required": False
        }
    
    # Check 30-day constraint
    sale_date = sale.get("created_at")
    if isinstance(sale_date, str):
        sale_date = datetime.fromisoformat(sale_date)
    
    days_old = (datetime.utcnow() - sale_date).days
    if days_old > 30:
        return {
            "can_delete": False,
            "reason": f"Transaction is {days_old} days old. Only transactions within 30 days can be deleted.",
            "is_admin_required": False,
            "days_old": days_old
        }
    
    # Check if user is admin
    is_admin = current_user.get("role") == "admin"
    
    return {
        "can_delete": is_admin,
        "reason": None if is_admin else "Only admin users can delete transactions",
        "is_admin_required": True,
        "days_old": days_old,
        "days_remaining": 30 - days_old
    }

# ===================== FARMER ROUTES =====================

@api_router.post("/farmers", response_model=Farmer)
async def create_farmer(farmer: FarmerCreate, current_user: dict = Depends(get_current_user)):
    farmer_obj = Farmer(**farmer.dict())
    await db.farmers.insert_one(farmer_obj.dict())
    return farmer_obj

@api_router.get("/farmers", response_model=List[Farmer])
async def get_farmers(current_user: dict = Depends(get_current_user)):
    farmers = await db.farmers.find({"is_active": True}).to_list(1000)
    return [Farmer(**f) for f in farmers]

@api_router.get("/farmers/{farmer_id}", response_model=Farmer)
async def get_farmer(farmer_id: str, current_user: dict = Depends(get_current_user)):
    farmer = await db.farmers.find_one({"id": farmer_id})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    return Farmer(**farmer)

@api_router.put("/farmers/{farmer_id}")
async def update_farmer(farmer_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    updates["updated_at"] = datetime.utcnow()
    result = await db.farmers.update_one({"id": farmer_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    return {"message": "Farmer updated successfully"}

@api_router.delete("/farmers/{farmer_id}")
async def delete_farmer(farmer_id: str, current_user: dict = Depends(require_admin)):
    result = await db.farmers.update_one(
        {"id": farmer_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    return {"message": "Farmer deactivated successfully"}

@api_router.get("/farmers/{farmer_id}/ledger")
async def get_farmer_ledger(farmer_id: str, current_user: dict = Depends(require_admin)):
    """Get farmer ledger with all transactions"""
    farmer = await db.farmers.find_one({"id": farmer_id})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    purchases = await db.farmer_purchases.find({"farmer_id": farmer_id}).sort("created_at", -1).to_list(1000)
    payments = await db.farmer_payments.find({"farmer_id": farmer_id}).sort("created_at", -1).to_list(1000)
    
    return {
        "farmer": Farmer(**farmer),
        "purchases": purchases,
        "payments": payments
    }

@api_router.post("/farmers/purchase", response_model=FarmerPurchase)
async def create_farmer_purchase(purchase: FarmerPurchaseCreate, current_user: dict = Depends(require_admin)):
    """Record produce purchase from farmer"""
    farmer = await db.farmers.find_one({"id": purchase.farmer_id})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    product = await db.products.find_one({"id": purchase.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    receipt_number = await generate_receipt_number()
    
    purchase_obj = FarmerPurchase(
        **purchase.dict(),
        farmer_name=farmer["name"],
        product_name=product["name"],
        receipt_number=receipt_number,
        created_by=current_user["id"]
    )
    await db.farmer_purchases.insert_one(purchase_obj.dict())
    
    # Update farmer ledger
    await db.farmers.update_one(
        {"id": purchase.farmer_id},
        {"$inc": {
            "total_supplied": purchase.quantity,
            "total_payable": purchase.total_amount,
            "total_paid": purchase.cash_amount + purchase.online_amount,
            "outstanding_dues": purchase.credit_amount
        }, "$set": {"updated_at": datetime.utcnow()}}
    )
    
    # Add to central stock (find or create central outlet stock)
    central_outlet = await db.outlets.find_one({"is_central": True})
    if central_outlet:
        existing_stock = await db.stock.find_one({
            "product_id": purchase.product_id,
            "outlet_id": central_outlet["id"]
        })
        
        if existing_stock:
            await db.stock.update_one(
                {"id": existing_stock["id"]},
                {"$inc": {
                    "quantity": purchase.quantity,
                    "stock_received": purchase.quantity
                }, "$set": {"updated_at": datetime.utcnow()}}
            )
        else:
            stock = Stock(
                product_id=purchase.product_id,
                outlet_id=central_outlet["id"],
                quantity=purchase.quantity,
                stock_received=purchase.quantity
            )
            await db.stock.insert_one(stock.dict())
        
        # Log movement
        movement = StockMovement(
            product_id=purchase.product_id,
            to_outlet_id=central_outlet["id"],
            quantity=purchase.quantity,
            movement_type="purchase",
            created_by=current_user["id"]
        )
        await db.stock_movements.insert_one(movement.dict())
    
    return purchase_obj

@api_router.get("/farmer-purchases")
async def get_farmer_purchases(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    farmer_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all farmer purchases with optional date filtering"""
    query = {}
    
    if farmer_id:
        query["farmer_id"] = farmer_id
    
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}
    
    purchases = await db.farmer_purchases.find(query).sort("created_at", -1).to_list(1000)
    # Convert MongoDB documents to dict without _id
    result = []
    for p in purchases:
        p.pop('_id', None)
        result.append(p)
    return result

@api_router.post("/farmer-purchases")
async def create_farmer_purchase_alt(purchase: FarmerPurchaseCreate, current_user: dict = Depends(get_current_user)):
    """Record produce purchase from farmer (alternative endpoint)"""
    
    # Handle manual farmer entry
    farmer_name = purchase.manual_farmer_name
    farmer_id = purchase.farmer_id
    if purchase.farmer_id != 'manual':
        farmer = await db.farmers.find_one({"id": purchase.farmer_id})
        if not farmer:
            raise HTTPException(status_code=404, detail="Farmer not found")
        farmer_name = farmer["name"]
    elif not purchase.manual_farmer_name:
        raise HTTPException(status_code=400, detail="Manual farmer name required")
    
    # Handle manual product entry
    product_name = purchase.manual_product_name
    product_unit = purchase.manual_product_unit or 'kg'
    product_id = purchase.product_id
    if purchase.product_id != 'manual':
        product = await db.products.find_one({"id": purchase.product_id})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        product_name = product["name"]
        product_unit = product.get("unit", "kg")
    elif not purchase.manual_product_name:
        raise HTTPException(status_code=400, detail="Manual product name required")
    
    # Validate outlet if provided
    outlet_id = purchase.outlet_id
    outlet_name = None
    if outlet_id:
        outlet = await db.outlets.find_one({"id": outlet_id})
        if outlet:
            outlet_name = outlet["name"]
    
    receipt_number = await generate_receipt_number()
    total = purchase.quantity * purchase.rate
    
    payment_mode = "cash" if purchase.payment_status == "paid" else "credit"
    
    purchase_obj = FarmerPurchase(
        farmer_id=farmer_id,
        farmer_name=farmer_name,
        product_id=product_id,
        product_name=product_name,
        quantity=purchase.quantity,
        rate=purchase.rate,
        total_amount=total,
        payment_mode=payment_mode,
        payment_status=purchase.payment_status,
        cash_amount=total if purchase.payment_status == "paid" else 0,
        credit_amount=total if purchase.payment_status == "credit" else 0,
        receipt_number=receipt_number,
        created_by=current_user["id"],
        outlet_id=outlet_id,
        outlet_name=outlet_name
    )
    await db.farmer_purchases.insert_one(purchase_obj.dict())
    
    # Update farmer ledger only if not manual entry
    if purchase.farmer_id != 'manual':
        paid_amount = purchase_obj.cash_amount + purchase_obj.online_amount
        await db.farmers.update_one(
            {"id": purchase.farmer_id},
            {"$inc": {
                "total_supplied": purchase.quantity,
                "total_payable": total,
                "total_paid": paid_amount,
                "outstanding_dues": purchase_obj.credit_amount
            }, "$set": {"updated_at": datetime.utcnow()}}
        )
    
    # UPDATE STOCK - Add to outlet's stock after procurement
    if outlet_id and product_id != 'manual':
        existing_stock = await db.stock.find_one({
            "product_id": product_id,
            "outlet_id": outlet_id
        })
        
        if existing_stock:
            await db.stock.update_one(
                {"id": existing_stock["id"]},
                {"$inc": {
                    "quantity": purchase.quantity,
                    "stock_received": purchase.quantity
                }, "$set": {"updated_at": datetime.utcnow()}}
            )
        else:
            # Create new stock record
            from uuid import uuid4
            stock_obj = {
                "id": str(uuid4()),
                "product_id": product_id,
                "outlet_id": outlet_id,
                "quantity": purchase.quantity,
                "opening_stock": purchase.quantity,
                "stock_received": purchase.quantity,
                "stock_sold": 0,
                "stock_damaged": 0,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            await db.stock.insert_one(stock_obj)
        
        # Record stock movement
        movement = {
            "id": str(uuid.uuid4()),
            "product_id": product_id,
            "from_outlet_id": None,
            "to_outlet_id": outlet_id,
            "quantity": purchase.quantity,
            "movement_type": "farmer_procurement",
            "reference_id": purchase_obj.id,
            "notes": f"Procurement from farmer: {farmer_name}",
            "created_by": current_user["id"],
            "created_at": datetime.utcnow()
        }
        await db.stock_movements.insert_one(movement)
    
    return {
        "message": "Purchase recorded successfully",
        "receipt_number": receipt_number,
        "total_amount": total,
        "stock_updated": outlet_id is not None and product_id != 'manual'
    }

@api_router.post("/farmers/payment")
async def record_farmer_payment(payment: FarmerPaymentCreate, current_user: dict = Depends(require_admin)):
    """Record payment to farmer"""
    farmer = await db.farmers.find_one({"id": payment.farmer_id})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    payment_obj = FarmerPayment(
        farmer_id=payment.farmer_id,
        amount=payment.amount,
        payment_mode=payment.payment_mode,
        notes=payment.notes,
        created_by=current_user["id"]
    )
    await db.farmer_payments.insert_one(payment_obj.dict())
    
    # Update farmer balance
    await db.farmers.update_one(
        {"id": payment.farmer_id},
        {"$inc": {
            "total_paid": payment.amount,
            "outstanding_dues": -payment.amount
        }, "$set": {"updated_at": datetime.utcnow()}}
    )
    
    return {"message": "Payment recorded successfully"}

# ===================== VENDOR ROUTES (KHATA/LEDGER SYSTEM) =====================

@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(vendor: VendorCreate, current_user: dict = Depends(get_current_user)):
    """Create a new vendor with auto-ledger creation"""
    vendor_obj = Vendor(**vendor.dict())
    await db.vendors.insert_one(vendor_obj.dict())
    return vendor_obj

@api_router.get("/vendors", response_model=List[Vendor])
async def get_vendors(
    search: Optional[str] = None,
    has_dues: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get vendors with search and filter support"""
    query = {"is_active": True}
    
    # Search by name, mobile, or village (case-insensitive partial match)
    if search:
        import re
        search_regex = re.compile(search, re.IGNORECASE)
        query["$or"] = [
            {"name": {"$regex": search_regex}},
            {"mobile": {"$regex": search_regex}},
            {"village": {"$regex": search_regex}}
        ]
    
    # Filter vendors with outstanding dues
    if has_dues:
        query["outstanding_dues"] = {"$gt": 0}
    
    vendors = await db.vendors.find(query).sort("name", 1).to_list(1000)
    return [Vendor(**v) for v in vendors]

@api_router.get("/vendors/search")
async def search_vendors(
    q: str,
    current_user: dict = Depends(get_current_user)
):
    """Quick search endpoint for vendors - used in autocomplete"""
    if not q or len(q.strip()) < 1:
        return []

    import re
    # Escape special regex chars so user input like "(", "+" doesn't crash the query
    safe_q = re.escape(q.strip())
    search_regex = {"$regex": safe_q, "$options": "i"}
    query = {
        "is_active": True,
        "$or": [
            {"name": search_regex},
            {"mobile": search_regex},
            {"village": search_regex},
            {"address": search_regex},
        ]
    }

    vendors = await db.vendors.find(query).limit(20).to_list(20)
    return [{
        "id": v["id"],
        "name": v["name"],
        "mobile": v.get("mobile", ""),
        "address": v.get("address", ""),
        "village": v.get("village", ""),
        "outstanding_dues": v.get("outstanding_dues", 0)
    } for v in vendors]


# ===================== GLOBAL SEARCH =====================

@api_router.get("/search/global")
async def global_search(
    q: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Unified search across customers, vendors, products, sales (bill#), and outlets.
    Returns results grouped by category, up to 10 per group.
    """
    if not q or len(q.strip()) < 1:
        return {
            "customers": [], "vendors": [], "products": [],
            "sales": [], "outlets": [], "total": 0
        }

    import re
    safe_q = re.escape(q.strip())
    rx = {"$regex": safe_q, "$options": "i"}

    # --- Customers ---
    customers_raw = await db.customers.find({
        "is_active": True,
        "$or": [
            {"name": rx}, {"name_hi": rx}, {"mobile": rx}, {"village": rx},
            {"address": rx}, {"folio_number": rx},
        ]
    }).limit(10).to_list(10)
    customers = [{
        "id": c["id"],
        "name": c.get("name", ""),
        "name_hi": c.get("name_hi", ""),
        "mobile": c.get("mobile", ""),
        "village": c.get("village", ""),
        "address": c.get("address", ""),
        "customer_type": c.get("customer_type", ""),
        "folio_number": c.get("folio_number", ""),
        "outstanding_balance": c.get("outstanding_balance", 0),
    } for c in customers_raw]

    # --- Vendors ---
    vendors_raw = await db.vendors.find({
        "is_active": True,
        "$or": [
            {"name": rx}, {"mobile": rx}, {"village": rx}, {"address": rx},
        ]
    }).limit(10).to_list(10)
    vendors = [{
        "id": v["id"],
        "name": v.get("name", ""),
        "mobile": v.get("mobile", ""),
        "village": v.get("village", ""),
        "address": v.get("address", ""),
        "outstanding_dues": v.get("outstanding_dues", 0),
    } for v in vendors_raw]

    # --- Products ---
    products_raw = await db.products.find({
        "is_active": True,
        "$or": [
            {"name": rx}, {"name_hi": rx},
            {"description": rx}, {"category": rx},
            {"varieties.name": rx}, {"varieties.name_hi": rx},
        ]
    }).limit(10).to_list(10)
    products = [{
        "id": p["id"],
        "name": p.get("name", ""),
        "name_hi": p.get("name_hi", ""),
        "unit": p.get("unit", ""),
        "category": p.get("category", ""),
        "varieties_count": len(p.get("varieties", []) or []),
    } for p in products_raw]

    # --- Sales (bill number, customer name) ---
    sales_raw = await db.sales.find({
        "$or": [
            {"bill_number": rx},
            {"customer_name": rx},
        ]
    }).sort("created_at", -1).limit(10).to_list(10)
    sales = [{
        "id": s["id"],
        "bill_number": s.get("bill_number", ""),
        "customer_name": s.get("customer_name", ""),
        "customer_id": s.get("customer_id"),
        "total_amount": s.get("total_amount", 0),
        "payment_mode": s.get("payment_mode", ""),
        "created_at": s.get("created_at").isoformat() if isinstance(s.get("created_at"), datetime) else str(s.get("created_at", "")),
        "is_deleted": s.get("is_deleted", False),
    } for s in sales_raw]

    # --- Outlets ---
    outlets_raw = await db.outlets.find({
        "is_active": True,
        "$or": [{"name": rx}, {"address": rx}, {"manager_name": rx}]
    }).limit(10).to_list(10)
    outlets = [{
        "id": o["id"],
        "name": o.get("name", ""),
        "address": o.get("address", ""),
        "manager_name": o.get("manager_name", ""),
        "outlet_type": o.get("outlet_type", ""),
    } for o in outlets_raw]

    total = len(customers) + len(vendors) + len(products) + len(sales) + len(outlets)

    return {
        "query": q.strip(),
        "customers": customers,
        "vendors": vendors,
        "products": products,
        "sales": sales,
        "outlets": outlets,
        "total": total,
    }


# ===================== REPORTS =====================

@api_router.get("/reports/transactions")
async def report_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    type: Optional[str] = "all",  # sale | purchase | all
    customer_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    outlet_id: Optional[str] = None,
    format: Optional[str] = "json",  # json | csv
    include_deleted: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """
    Unified transaction report. Filters: date range, type, customer/vendor, outlet.
    Output: JSON list or CSV stream.
    Each row: date, time, type, ref, person_name, outlet, product, variety, qty, rate, total, payment_mode, is_cancelled.
    """
    from fastapi.responses import StreamingResponse
    import io
    import csv

    # ---- date filters ----
    date_q: dict = {}
    if start_date:
        try:
            date_q["$gte"] = datetime.fromisoformat(start_date)
        except Exception:
            pass
    if end_date:
        try:
            date_q["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        except Exception:
            pass

    rows = []

    # ---- agent restriction ----
    is_agent = current_user.get("role") == "agent"
    forced_outlet = current_user.get("outlet_id") if is_agent else None
    eff_outlet = forced_outlet or outlet_id

    # ---- SALES ----
    if type in (None, "all", "sale"):
        sale_query: dict = {}
        if not include_deleted:
            sale_query["$or"] = [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]
        if date_q:
            sale_query["created_at"] = date_q
        if customer_id:
            sale_query["customer_id"] = customer_id
        if eff_outlet:
            sale_query["outlet_id"] = eff_outlet

        sales_docs = await db.sales.find(sale_query).sort("created_at", -1).to_list(2000)

        outlet_ids = list({s.get("outlet_id") for s in sales_docs if s.get("outlet_id")})
        outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(500) if outlet_ids else []
        outlets_map = {o["id"]: o.get("name", "") for o in outlets_list}

        for s in sales_docs:
            dt = s.get("created_at") or datetime.utcnow()
            outlet_name = outlets_map.get(s.get("outlet_id"), "")
            person = s.get("customer_name") or "-"
            ref = s.get("bill_number", "")
            payment = s.get("payment_mode", "")
            is_cancelled = bool(s.get("is_deleted"))
            for it in (s.get("items") or []):
                rows.append({
                    "date": dt.strftime("%Y-%m-%d") if isinstance(dt, datetime) else str(dt),
                    "time": dt.strftime("%H:%M:%S") if isinstance(dt, datetime) else "",
                    "type": "Sale",
                    "reference": ref,
                    "person_name": person,
                    "outlet": outlet_name,
                    "product": it.get("product_name", ""),
                    "variety": it.get("variety_name", "") or "",
                    "quantity": it.get("quantity", 0),
                    "rate": it.get("rate", 0),
                    "amount": it.get("amount", 0),
                    "total": s.get("total_amount", 0),
                    "payment_mode": payment,
                    "is_cancelled": is_cancelled,
                })

    # ---- PURCHASES (vendor procurement) ----
    # Skip purchases entirely when filtering by customer (purchases are vendor-scoped)
    skip_purchases = bool(customer_id)
    if (type in (None, "all", "purchase")) and not skip_purchases:
        proc_query: dict = {}
        if not include_deleted:
            proc_query["$or"] = [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]
        if date_q:
            proc_query["created_at"] = date_q
        if vendor_id:
            proc_query["vendor_id"] = vendor_id
        if eff_outlet:
            proc_query["outlet_id"] = eff_outlet

        proc_docs = await db.vendor_procurement.find(proc_query).sort("created_at", -1).to_list(2000)

        outlet_ids = list({p.get("outlet_id") for p in proc_docs if p.get("outlet_id")})
        outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(500) if outlet_ids else []
        outlets_map = {o["id"]: o.get("name", "") for o in outlets_list}

        for p in proc_docs:
            dt = p.get("created_at") or datetime.utcnow()
            outlet_name = outlets_map.get(p.get("outlet_id"), "")
            rows.append({
                "date": dt.strftime("%Y-%m-%d") if isinstance(dt, datetime) else str(dt),
                "time": dt.strftime("%H:%M:%S") if isinstance(dt, datetime) else "",
                "type": "Purchase",
                "reference": p.get("receipt_number", ""),
                "person_name": p.get("vendor_name") or p.get("manual_vendor_name") or "-",
                "outlet": outlet_name,
                "product": p.get("product_name") or p.get("manual_product_name") or "",
                "variety": p.get("variety_name") or "",
                "quantity": p.get("quantity", 0),
                "rate": p.get("rate", 0),
                "amount": p.get("total_amount", 0),
                "total": p.get("total_amount", 0),
                "payment_mode": p.get("payment_mode", ""),
                "is_cancelled": bool(p.get("is_deleted")),
            })

    # newest first
    rows.sort(key=lambda r: (r.get("date", ""), r.get("time", "")), reverse=True)

    if (format or "json").lower() == "csv":
        buffer = io.StringIO()
        fieldnames = [
            "date", "time", "type", "reference", "person_name", "outlet",
            "product", "variety", "quantity", "rate", "amount", "total",
            "payment_mode", "is_cancelled",
        ]
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        buffer.seek(0)
        filename = f"transactions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {"count": len(rows), "rows": rows}


@api_router.get("/reports/raw")
async def report_raw(
    type: str = "sale",  # sale | purchase
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: Optional[str] = "csv",
    current_user: dict = Depends(require_admin),
):
    """
    Raw transaction dump for admins (full payload). Includes deleted entries with the deletion fields.
    """
    from fastapi.responses import StreamingResponse
    import io
    import csv
    import json

    date_q: dict = {}
    if start_date:
        try:
            date_q["$gte"] = datetime.fromisoformat(start_date)
        except Exception:
            pass
    if end_date:
        try:
            date_q["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        except Exception:
            pass

    coll = db.sales if type == "sale" else db.vendor_procurement
    query: dict = {}
    if date_q:
        query["created_at"] = date_q

    docs = await coll.find(query).sort("created_at", -1).to_list(5000)
    for d in docs:
        d.pop("_id", None)

    if (format or "csv").lower() == "csv":
        buffer = io.StringIO()
        # union of all keys for header
        keys = set()
        for d in docs:
            keys.update(d.keys())
        fieldnames = sorted(keys)
        writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for d in docs:
            row = {}
            for k in fieldnames:
                v = d.get(k)
                if isinstance(v, (list, dict)):
                    row[k] = json.dumps(v, default=str)
                elif isinstance(v, datetime):
                    row[k] = v.isoformat()
                else:
                    row[k] = v
            writer.writerow(row)
        buffer.seek(0)
        filename = f"raw_{type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {"count": len(docs), "rows": docs}


@api_router.get("/vendors/with-dues")
async def get_vendors_with_dues(current_user: dict = Depends(get_current_user)):
    """Get all vendors with outstanding dues (for Khata overview)"""
    vendors = await db.vendors.find({
        "is_active": True,
        "outstanding_dues": {"$gt": 0}
    }).sort("outstanding_dues", -1).to_list(1000)
    
    total_dues = sum(v.get("outstanding_dues", 0) for v in vendors)
    
    return {
        "vendors": [Vendor(**v) for v in vendors],
        "total_dues": total_dues,
        "count": len(vendors)
    }

@api_router.get("/vendors/{vendor_id}", response_model=Vendor)
async def get_vendor(vendor_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific vendor"""
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return Vendor(**vendor)

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update a vendor"""
    updates["updated_at"] = datetime.utcnow()
    result = await db.vendors.update_one({"id": vendor_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor updated successfully"}

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, current_user: dict = Depends(require_admin)):
    """Soft delete a vendor"""
    result = await db.vendors.update_one(
        {"id": vendor_id},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor deactivated successfully"}

@api_router.get("/vendors/{vendor_id}/ledger")
async def get_vendor_ledger(
    vendor_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive vendor ledger (Khata) with all transactions"""
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Build date filter
    date_filter = {}
    if start_date:
        date_filter["$gte"] = datetime.fromisoformat(start_date)
    if end_date:
        date_filter["$lte"] = datetime.fromisoformat(end_date)
    
    # Get purchases from this vendor
    purchases_query = {"vendor_id": vendor_id}
    if date_filter:
        purchases_query["created_at"] = date_filter
    
    purchases_raw = await db.vendor_procurement.find(purchases_query).sort("created_at", -1).to_list(1000)
    
    # Get payments to this vendor
    payments_query = {"vendor_id": vendor_id}
    if date_filter:
        payments_query["created_at"] = date_filter
    
    payments_raw = await db.vendor_payments.find(payments_query).sort("created_at", -1).to_list(1000)
    
    # Clean MongoDB _id from results
    def clean_doc(doc):
        doc.pop('_id', None)
        return doc
    
    purchases = [clean_doc(p) for p in purchases_raw]
    payments = [clean_doc(p) for p in payments_raw]

    # Split active vs cancelled — cancelled are shown but excluded from totals
    active_purchases = [p for p in purchases if not p.get("is_deleted", False)]

    # Calculate ledger summary
    total_purchases = sum(p.get("total_amount", 0) for p in active_purchases)
    total_credit_given = sum(p.get("credit_amount", 0) for p in active_purchases)
    total_payments = sum(p.get("amount", 0) for p in payments)

    # Build transaction timeline (combined purchases and payments sorted by date)
    transactions = []
    for p in purchases:
        is_cancelled = p.get("is_deleted", False)
        transactions.append({
            "id": p["id"],
            "type": "purchase",
            "date": p["created_at"],
            "reference": p.get("receipt_number", ""),
            "description": f"Purchase - {p.get('product_name', 'N/A')}" + (" (CANCELLED)" if is_cancelled else ""),
            "debit": 0 if is_cancelled else p.get("credit_amount", 0),
            "credit": 0,
            "quantity": p.get("quantity", 0),
            "rate": p.get("rate", 0),
            "total_amount": p.get("total_amount", 0),
            "payment_mode": p.get("payment_mode", ""),
            "is_cancelled": is_cancelled,
            "deleted_at": p.get("deleted_at"),
            "deletion_reason": p.get("deletion_reason", ""),
        })
    
    for p in payments:
        transactions.append({
            "id": p["id"],
            "type": "payment",
            "date": p["created_at"],
            "reference": "",
            "description": f"Payment - {p.get('payment_mode', 'cash').upper()}",
            "debit": 0,
            "credit": p.get("amount", 0),  # Payment reduces dues
            "payment_mode": p.get("payment_mode", "cash"),
            "notes": p.get("notes", "")
        })
    
    # Sort by date descending
    transactions.sort(key=lambda x: x["date"], reverse=True)
    
    # Remove _id from vendor
    vendor.pop('_id', None)
    
    return {
        "vendor": Vendor(**vendor),
        "transactions": transactions,
        "summary": {
            "total_transactions": len(purchases),
            "total_purchases": total_purchases,
            "total_credit_given": total_credit_given,
            "total_payments": total_payments,
            "outstanding_dues": vendor.get("outstanding_dues", 0)
        },
        "purchases": purchases,
        "payments": payments
    }

@api_router.post("/vendors/payment")
async def record_vendor_payment(payment: VendorPaymentCreate, current_user: dict = Depends(get_current_user)):
    """Record a payment to vendor - adjusts dues automatically"""
    vendor = await db.vendors.find_one({"id": payment.vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    payment_obj = VendorPayment(
        vendor_id=payment.vendor_id,
        amount=payment.amount,
        payment_mode=payment.payment_mode,
        notes=payment.notes,
        created_by=current_user["id"]
    )
    await db.vendor_payments.insert_one(payment_obj.dict())
    
    # Update vendor balance - auto-adjust dues
    new_dues = max(0, vendor.get("outstanding_dues", 0) - payment.amount)
    await db.vendors.update_one(
        {"id": payment.vendor_id},
        {
            "$inc": {"total_paid": payment.amount},
            "$set": {
                "outstanding_dues": new_dues,
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    return {
        "message": "Payment recorded successfully",
        "new_dues": new_dues
    }

# ===================== DASHBOARD & REPORTS =====================

@api_router.get("/dashboard")
async def get_dashboard(current_user: dict = Depends(get_current_user)):
    """Get dashboard data"""
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    
    query = {}
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    
    # Today's sales
    today_query = {**query, "created_at": {"$gte": today}}
    today_sales = await db.sales.find(today_query).to_list(1000)
    today_total = sum(s.get("total_amount", 0) for s in today_sales)
    today_cash = sum(s.get("cash_amount", 0) for s in today_sales)
    today_credit = sum(s.get("credit_amount", 0) for s in today_sales)
    
    # Month's sales
    month_query = {**query, "created_at": {"$gte": month_start}}
    month_sales = await db.sales.find(month_query).to_list(1000)
    month_total = sum(s.get("total_amount", 0) for s in month_sales)
    month_cash = sum(s.get("cash_amount", 0) for s in month_sales)
    month_credit = sum(s.get("credit_amount", 0) for s in month_sales)
    
    # Stock summary
    stock_query = {}
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        stock_query["outlet_id"] = current_user["outlet_id"]
    
    stocks = await db.stock.find(stock_query).to_list(1000)
    low_stock_items = [s for s in stocks if s.get("quantity", 0) < 10]
    
    # Counts
    total_customers = await db.customers.count_documents({"is_active": True})
    total_farmers = await db.farmers.count_documents({"is_active": True})
    total_products = await db.products.count_documents({"is_active": True})
    total_outlets = await db.outlets.count_documents({"is_active": True})
    
    # Outstanding balances
    customers = await db.customers.find({"outstanding_balance": {"$gt": 0}}).to_list(1000)
    total_customer_outstanding = sum(c.get("outstanding_balance", 0) for c in customers)
    
    farmers = await db.farmers.find({"outstanding_dues": {"$gt": 0}}).to_list(1000)
    total_farmer_dues = sum(f.get("outstanding_dues", 0) for f in farmers)
    
    return {
        "today": {
            "sales_count": len(today_sales),
            "total": today_total,
            "cash": today_cash,
            "credit": today_credit,
            "online": today_total - today_cash - today_credit
        },
        "month": {
            "sales_count": len(month_sales),
            "total": month_total,
            "cash": month_cash,
            "credit": month_credit,
            "online": month_total - month_cash - month_credit
        },
        "counts": {
            "customers": total_customers,
            "farmers": total_farmers,
            "products": total_products,
            "outlets": total_outlets
        },
        "outstanding": {
            "customer_dues": total_customer_outstanding,
            "farmer_dues": total_farmer_dues
        },
        "low_stock_count": len(low_stock_items)
    }

@api_router.get("/reports/sales")
async def get_sales_report(
    outlet_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get sales report with filters"""
    query = {}
    
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}
    
    sales = await db.sales.find(query).sort("created_at", -1).to_list(1000)
    
    # Remove MongoDB _id from all sales
    for sale in sales:
        sale.pop('_id', None)
    
    # Summary
    total_amount = sum(s.get("total_amount", 0) for s in sales)
    total_cash = sum(s.get("cash_amount", 0) for s in sales)
    total_online = sum(s.get("online_amount", 0) for s in sales)
    total_credit = sum(s.get("credit_amount", 0) for s in sales)
    
    # Batch fetch outlets to avoid N+1 query
    outlet_ids = list(set(s.get("outlet_id") for s in sales if s.get("outlet_id")))
    outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(1000) if outlet_ids else []
    outlets_map = {o["id"]: o for o in outlets_list}
    
    # Enrich with outlet names
    for sale in sales:
        outlet = outlets_map.get(sale.get("outlet_id"))
        sale["outlet_name"] = outlet["name"] if outlet else "Unknown"
    
    return {
        "sales": sales,
        "summary": {
            "count": len(sales),
            "total_amount": total_amount,
            "cash": total_cash,
            "online": total_online,
            "credit": total_credit
        }
    }

@api_router.get("/reports/stock")
async def get_stock_report(outlet_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get stock report"""
    query = {}
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    stocks = await db.stock.find(query).to_list(1000)
    
    # Batch fetch products and outlets to avoid N+1 query
    product_ids = list(set(s.get("product_id") for s in stocks if s.get("product_id")))
    outlet_ids = list(set(s.get("outlet_id") for s in stocks if s.get("outlet_id")))
    
    products_list = await db.products.find({"id": {"$in": product_ids}}).to_list(1000) if product_ids else []
    outlets_list = await db.outlets.find({"id": {"$in": outlet_ids}}).to_list(1000) if outlet_ids else []
    
    products_map = {p["id"]: p for p in products_list}
    outlets_map = {o["id"]: o for o in outlets_list}
    
    result = []
    for s in stocks:
        s.pop('_id', None)  # Remove MongoDB _id
        product = products_map.get(s.get("product_id"))
        outlet = outlets_map.get(s.get("outlet_id"))
        result.append({
            **s,
            "product_name": product["name"] if product else "Unknown",
            "product_unit": product["unit"] if product else "",
            "outlet_name": outlet["name"] if outlet else "Unknown"
        })
    
    return result

@api_router.get("/reports/customers/export")
async def export_customer_ledger(customer_id: Optional[str] = None, current_user: dict = Depends(require_admin)):
    """Export customer ledger to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    
    # Headers
    headers = ["Customer Name", "Mobile", "Village", "Total Purchases", "Total Credit", "Total Paid", "Outstanding"]
    ws.append(headers)
    
    query = {"is_active": True}
    if customer_id:
        query["id"] = customer_id
    
    customers = await db.customers.find(query).to_list(1000)
    
    for c in customers:
        ws.append([
            c.get("name", ""),
            c.get("mobile", ""),
            c.get("village", ""),
            c.get("total_purchases", 0),
            c.get("total_credit", 0),
            c.get("total_paid", 0),
            c.get("outstanding_balance", 0)
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_ledger.xlsx"}
    )

@api_router.get("/reports/farmers/export")
async def export_farmer_ledger(farmer_id: Optional[str] = None, current_user: dict = Depends(require_admin)):
    """Export farmer ledger to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Farmer Ledger"
    
    # Headers
    headers = ["Farmer Name", "Village", "Mobile", "Member", "Total Supplied", "Total Payable", "Total Paid", "Outstanding Dues"]
    ws.append(headers)
    
    query = {"is_active": True}
    if farmer_id:
        query["id"] = farmer_id
    
    farmers = await db.farmers.find(query).to_list(1000)
    
    for f in farmers:
        ws.append([
            f.get("name", ""),
            f.get("village", ""),
            f.get("mobile", ""),
            "Yes" if f.get("is_member") else "No",
            f.get("total_supplied", 0),
            f.get("total_payable", 0),
            f.get("total_paid", 0),
            f.get("outstanding_dues", 0)
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=farmer_ledger.xlsx"}
    )

@api_router.get("/reports/sales/export")
async def export_sales_report(
    outlet_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export sales report to Excel"""
    query = {}
    
    if current_user["role"] == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]
    elif outlet_id:
        query["outlet_id"] = outlet_id
    
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}
    
    sales = await db.sales.find(query).sort("created_at", -1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ["Bill No", "Date", "Outlet", "Customer", "Total", "Cash", "Online", "Credit", "Payment Mode"]
    ws.append(headers)
    
    for s in sales:
        outlet = await db.outlets.find_one({"id": s["outlet_id"]})
        ws.append([
            s.get("bill_number", ""),
            s.get("created_at", "").strftime("%Y-%m-%d %H:%M") if isinstance(s.get("created_at"), datetime) else str(s.get("created_at", "")),
            outlet["name"] if outlet else "Unknown",
            s.get("customer_name", "Walk-in"),
            s.get("total_amount", 0),
            s.get("cash_amount", 0),
            s.get("online_amount", 0),
            s.get("credit_amount", 0),
            s.get("payment_mode", "")
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
    )

# ===================== NOTIFICATIONS =====================

async def create_notification(user_id: str, title: str, message: str, notif_type: str, ref_type: str = None, ref_id: str = None):
    """Helper function to create notifications"""
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        type=notif_type,
        reference_type=ref_type,
        reference_id=ref_id
    )
    await db.notifications.insert_one(notification.dict())
    return notification

async def notify_admins(title: str, message: str, notif_type: str, ref_type: str = None, ref_id: str = None):
    """Send notification to all admins"""
    admins = await db.users.find({"role": "admin", "is_active": True}).to_list(100)
    for admin in admins:
        await create_notification(admin["id"], title, message, notif_type, ref_type, ref_id)

async def notify_outlet_agents(outlet_id: str, title: str, message: str, notif_type: str, ref_type: str = None, ref_id: str = None):
    """Send notification to agents of a specific outlet"""
    agents = await db.users.find({"role": "agent", "outlet_id": outlet_id, "is_active": True}).to_list(100)
    for agent in agents:
        await create_notification(agent["id"], title, message, notif_type, ref_type, ref_id)

@api_router.get("/notifications")
async def get_notifications(unread_only: bool = False, current_user: dict = Depends(get_current_user)):
    """Get notifications for current user"""
    query = {"user_id": current_user["id"]}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(query).sort("created_at", -1).to_list(100)
    result = []
    for n in notifications:
        n.pop('_id', None)
        result.append(n)
    return result

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: dict = Depends(get_current_user)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({"user_id": current_user["id"], "is_read": False})
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["id"]},
        {"$set": {"is_read": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all notifications as read"""
    await db.notifications.update_many(
        {"user_id": current_user["id"], "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "All notifications marked as read"}

# ===================== SHAREHOLDER UPGRADE =====================

@api_router.post("/shareholder-upgrade/request")
async def request_shareholder_upgrade(
    folio_number: Optional[str] = None,
    share_value: Optional[float] = None,
    certificate_data: Optional[str] = None,
    certificate_filename: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Request upgrade to shareholder status"""
    if current_user["role"] != "farmer":
        raise HTTPException(status_code=403, detail="Only farmers can request shareholder upgrade")
    
    # Check if already a shareholder
    user = await db.users.find_one({"id": current_user["id"]})
    if user and user.get("is_shareholder"):
        raise HTTPException(status_code=400, detail="You are already a shareholder")
    
    # Check if there's a pending request
    existing = await db.shareholder_upgrades.find_one({
        "user_id": current_user["id"],
        "status": "pending"
    })
    if existing:
        raise HTTPException(status_code=400, detail="You already have a pending upgrade request")
    
    upgrade_request = ShareholderUpgradeRequest(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", current_user.get("username")),
        folio_number=folio_number,
        share_value=share_value,
        certificate_data=certificate_data,
        certificate_filename=certificate_filename,
        status="pending"
    )
    
    await db.shareholder_upgrades.insert_one(upgrade_request.dict())
    
    # Notify admins
    await notify_admins(
        title="Shareholder Upgrade Request",
        message=f"{upgrade_request.user_name} has requested shareholder upgrade",
        notif_type="shareholder",
        ref_type="shareholder_upgrade",
        ref_id=upgrade_request.id
    )
    
    return {"message": "Upgrade request submitted successfully", "request_id": upgrade_request.id}

@api_router.get("/shareholder-upgrade/requests")
async def get_shareholder_requests(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get shareholder upgrade requests - Admin sees all, farmers see their own"""
    query = {}
    
    if current_user["role"] == "farmer":
        query["user_id"] = current_user["id"]
    elif current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if status:
        query["status"] = status
    
    requests = await db.shareholder_upgrades.find(query).sort("created_at", -1).to_list(100)
    result = []
    for r in requests:
        r.pop('_id', None)
        result.append(r)
    return result

@api_router.get("/shareholder-upgrade/pending-count")
async def get_pending_shareholder_count(current_user: dict = Depends(require_admin)):
    """Get count of pending shareholder requests"""
    count = await db.shareholder_upgrades.count_documents({"status": "pending"})
    return {"count": count}

@api_router.put("/shareholder-upgrade/{request_id}/approve")
async def approve_shareholder_upgrade(request_id: str, remark: Optional[str] = None, current_user: dict = Depends(require_admin)):
    """Approve a shareholder upgrade request"""
    upgrade = await db.shareholder_upgrades.find_one({"id": request_id})
    if not upgrade:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if upgrade["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Update the upgrade request
    await db.shareholder_upgrades.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "admin_remark": remark,
            "updated_at": datetime.utcnow()
        }}
    )
    
    # Update the user to be a shareholder with folio number and share value
    user_update = {
        "is_shareholder": True, 
        "updated_at": datetime.utcnow()
    }
    if upgrade.get("folio_number"):
        user_update["folio_number"] = upgrade["folio_number"]
    if upgrade.get("share_value"):
        user_update["share_value"] = upgrade["share_value"]
    
    await db.users.update_one(
        {"id": upgrade["user_id"]},
        {"$set": user_update}
    )
    
    # Also update farmer record if exists
    await db.farmers.update_one(
        {"user_id": upgrade["user_id"]},
        {"$set": user_update}
    )
    
    # Notify the farmer
    await create_notification(
        upgrade["user_id"],
        "Shareholder Status Approved!",
        "Congratulations! Your shareholder upgrade request has been approved.",
        "shareholder",
        "shareholder_upgrade",
        request_id
    )
    
    return {"message": "Shareholder upgrade approved"}

@api_router.put("/shareholder-upgrade/{request_id}/reject")
async def reject_shareholder_upgrade(request_id: str, remark: str = "Request rejected", current_user: dict = Depends(require_admin)):
    """Reject a shareholder upgrade request"""
    upgrade = await db.shareholder_upgrades.find_one({"id": request_id})
    if not upgrade:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if upgrade["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    await db.shareholder_upgrades.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["id"],
            "admin_remark": remark,
            "updated_at": datetime.utcnow()
        }}
    )
    
    # Notify the farmer
    await create_notification(
        upgrade["user_id"],
        "Shareholder Request Rejected",
        f"Your shareholder upgrade request was rejected. Reason: {remark}",
        "shareholder",
        "shareholder_upgrade",
        request_id
    )
    
    return {"message": "Shareholder upgrade rejected"}

# ===================== VENDOR PROCUREMENT =====================

@api_router.post("/vendor-procurement")
async def create_vendor_procurement(procurement: VendorProcurementBase, current_user: dict = Depends(get_current_user)):
    """Record a purchase from a vendor"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Only admin or agent can record vendor procurement")
    
    # Handle manual vendor entry
    vendor_name = procurement.manual_vendor_name
    vendor_id = procurement.vendor_id
    if procurement.vendor_id != 'manual':
        vendor = await db.vendors.find_one({"id": procurement.vendor_id, "is_active": True})
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        vendor_name = vendor["name"]
    elif not procurement.manual_vendor_name:
        raise HTTPException(status_code=400, detail="Manual vendor name required")
    
    # Handle manual product entry
    product_name = procurement.manual_product_name
    product_unit = procurement.manual_product_unit or 'kg'
    product_id = procurement.product_id
    if procurement.product_id != 'manual':
        product = await db.products.find_one({"id": procurement.product_id, "is_active": True})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        product_name = product["name"]
        product_unit = product.get("unit", "kg")
    elif not procurement.manual_product_name:
        raise HTTPException(status_code=400, detail="Manual product name required")
    
    # Validate outlet
    outlet = await db.outlets.find_one({"id": procurement.outlet_id, "is_active": True})
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    outlet_name = outlet["name"]
    
    # Calculate total and credit
    total_amount = procurement.quantity * procurement.rate
    paid_amount = procurement.cash_amount + procurement.online_amount
    credit_amount = total_amount - paid_amount
    payment_status = "paid" if credit_amount <= 0 else "credit"
    
    # Generate receipt number
    count = await db.vendor_procurement.count_documents({})
    receipt_number = f"VP{datetime.now().strftime('%Y%m%d')}{str(count + 1).zfill(4)}"
    
    # Create procurement record
    proc_obj = VendorProcurement(
        vendor_id=vendor_id,
        vendor_name=vendor_name,
        product_id=product_id,
        product_name=product_name,
        outlet_id=procurement.outlet_id,
        outlet_name=outlet_name,
        quantity=procurement.quantity,
        rate=procurement.rate,
        payment_mode=procurement.payment_mode,
        cash_amount=procurement.cash_amount,
        online_amount=procurement.online_amount,
        notes=procurement.notes,
        receipt_number=receipt_number,
        total_amount=total_amount,
        credit_amount=credit_amount,
        payment_status=payment_status,
        created_by=current_user["id"]
    )
    
    await db.vendor_procurement.insert_one(proc_obj.dict())
    
    # Add stock to the outlet (only if product is not manual)
    if product_id != 'manual':
        existing_stock = await db.stock.find_one({
            "product_id": product_id,
            "outlet_id": procurement.outlet_id
        })
        
        if existing_stock:
            await db.stock.update_one(
                {"id": existing_stock["id"]},
                {"$inc": {
                    "quantity": procurement.quantity,
                    "stock_received": procurement.quantity
                }, "$set": {"updated_at": datetime.utcnow()}}
            )
        else:
            new_stock = Stock(
                product_id=product_id,
                outlet_id=procurement.outlet_id,
                quantity=procurement.quantity,
                stock_received=procurement.quantity
            )
            await db.stock.insert_one(new_stock.dict())
        
        # Log stock movement
        movement = StockMovement(
            product_id=product_id,
            to_outlet_id=procurement.outlet_id,
            quantity=procurement.quantity,
            movement_type="vendor_procurement",
            reason=f"Vendor procurement from {vendor_name}. Receipt: {receipt_number}",
            created_by=current_user["id"]
        )
        await db.stock_movements.insert_one(movement.dict())
    
    # Update vendor ledger only if not manual entry
    if vendor_id != 'manual':
        await db.vendors.update_one(
            {"id": vendor_id},
            {
                "$inc": {
                    "total_purchases": total_amount,
                    "total_paid": paid_amount,
                    "outstanding_dues": credit_amount,
                    "transaction_count": 1
                }, 
                "$set": {
                    "updated_at": datetime.utcnow(),
                    "last_transaction_date": datetime.utcnow()
                }
            }
        )
    
    return {
        "message": "Vendor procurement recorded successfully",
        "receipt_number": receipt_number,
        "procurement_id": proc_obj.id,
        "stock_updated": product_id != 'manual'
    }

@api_router.get("/vendor-procurement")
async def get_vendor_procurement(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_deleted: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get vendor procurement records (agent restricted to own outlet)"""
    query = {}

    # Hide cancelled by default
    if not include_deleted:
        query["$or"] = [{"is_deleted": {"$exists": False}}, {"is_deleted": False}]

    # Agent-only outlet restriction
    if current_user.get("role") == "agent" and current_user.get("outlet_id"):
        query["outlet_id"] = current_user["outlet_id"]

    if vendor_id:
        query["vendor_id"] = vendor_id

    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query and isinstance(query["created_at"], dict):
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date) + timedelta(days=1)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date) + timedelta(days=1)}

    procurements = await db.vendor_procurement.find(query).sort("created_at", -1).to_list(500)
    result = []
    for p in procurements:
        p.pop('_id', None)
        result.append(p)
    return result

@api_router.get("/vendor-procurement/{procurement_id}")
async def get_procurement_detail(procurement_id: str, current_user: dict = Depends(get_current_user)):
    """Get details of a specific procurement"""
    procurement = await db.vendor_procurement.find_one({"id": procurement_id})
    if not procurement:
        raise HTTPException(status_code=404, detail="Procurement not found")
    procurement.pop('_id', None)
    return procurement


@api_router.delete("/vendor-procurement/{procurement_id}")
async def delete_vendor_procurement(
    procurement_id: str,
    reason: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    """
    Delete a vendor procurement with automatic reversal of:
    1. Inventory/Stock - subtract the received quantity from the outlet
    2. Vendor Ledger - reverse outstanding_dues, totals, count
    Constraints: admin only, 30-day limit, soft-delete with audit log.
    """
    procurement = await db.vendor_procurement.find_one({"id": procurement_id})
    if not procurement:
        raise HTTPException(status_code=404, detail="Procurement not found")

    if procurement.get("is_deleted", False):
        raise HTTPException(status_code=400, detail="Transaction already deleted")

    proc_date = procurement.get("created_at")
    if isinstance(proc_date, str):
        proc_date = datetime.fromisoformat(proc_date)
    days_old = (datetime.utcnow() - proc_date).days
    if days_old > 30:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete transaction older than 30 days. This transaction is {days_old} days old."
        )

    reversal_details = {
        "stock_reversed": False,
        "vendor_ledger_adjusted": False,
        "vendor_id": None,
        "dues_reversed": 0,
        "paid_amount_reversed": 0,
    }

    outlet_id = procurement.get("outlet_id")
    product_id = procurement.get("product_id")
    quantity = procurement.get("quantity", 0)

    # 1. REVERSE STOCK (subtract received quantity at that outlet)
    if product_id and product_id != "manual" and quantity > 0 and outlet_id:
        stock = await db.stock.find_one({"outlet_id": outlet_id, "product_id": product_id})
        if stock:
            new_qty = max(0, (stock.get("quantity") or 0) - quantity)
            new_received = max(0, (stock.get("stock_received") or 0) - quantity)
            await db.stock.update_one(
                {"id": stock["id"]},
                {
                    "$set": {
                        "quantity": new_qty,
                        "stock_received": new_received,
                        "updated_at": datetime.utcnow(),
                    }
                },
            )
            reversal_details["stock_reversed"] = True

    # 2. REVERSE VENDOR LEDGER
    vendor_id = procurement.get("vendor_id")
    total_amount = procurement.get("total_amount", 0)
    credit_amount = procurement.get("credit_amount", 0)
    cash_amount = procurement.get("cash_amount", 0)
    online_amount = procurement.get("online_amount", 0)
    paid_amount = cash_amount + online_amount

    if vendor_id and vendor_id != "manual":
        vendor = await db.vendors.find_one({"id": vendor_id})
        if vendor:
            update_ops = {
                "$inc": {
                    "total_purchases": -total_amount,
                    "transaction_count": -1,
                },
                "$set": {"updated_at": datetime.utcnow()},
            }
            if credit_amount > 0:
                update_ops["$inc"]["outstanding_dues"] = -credit_amount
                reversal_details["dues_reversed"] = credit_amount
            if paid_amount > 0:
                update_ops["$inc"]["total_paid"] = -paid_amount
                reversal_details["paid_amount_reversed"] = paid_amount
            await db.vendors.update_one({"id": vendor_id}, update_ops)
            reversal_details["vendor_ledger_adjusted"] = True
            reversal_details["vendor_id"] = vendor_id

    # 3. SOFT DELETE
    await db.vendor_procurement.update_one(
        {"id": procurement_id},
        {
            "$set": {
                "is_deleted": True,
                "deleted_at": datetime.utcnow(),
                "deleted_by": current_user["id"],
                "deletion_reason": reason,
            }
        },
    )

    # 4. AUDIT LOG
    admin = await db.users.find_one({"id": current_user["id"]})
    admin_name = admin.get("full_name", admin.get("username", "Admin")) if admin else "Admin"

    procurement.pop("_id", None)
    try:
        audit_log = TransactionDeletionLog(
            transaction_type="vendor_procurement",
            transaction_id=procurement_id,
            original_data=procurement,
            bill_number=procurement.get("receipt_number"),
            customer_id=vendor_id,
            customer_name=procurement.get("vendor_name"),
            total_amount=total_amount,
            deleted_by=current_user["id"],
            deleted_by_name=admin_name,
            deletion_reason=reason,
            reversal_details=reversal_details,
        )
        await db.transaction_deletion_logs.insert_one(audit_log.dict())
    except Exception as e:
        logging.error(f"Failed writing audit log for vendor procurement deletion: {e}")

    return {
        "message": "Vendor procurement deleted successfully",
        "procurement_id": procurement_id,
        "receipt_number": procurement.get("receipt_number"),
        "reversal_details": reversal_details,
    }


# ===================== INITIALIZATION =====================

@api_router.post("/init/setup")
async def initial_setup():
    """Initialize the system with default admin and central outlet"""
    # Check if already initialized
    admin = await db.users.find_one({"role": "admin"})
    if admin:
        return {"message": "System already initialized", "admin_exists": True}
    
    # Create default admin
    admin_user = UserInDB(
        username="admin",
        full_name="System Administrator",
        role="admin",
        hashed_password=get_password_hash("admin123"),
        is_active=True
    )
    await db.users.insert_one(admin_user.dict())
    
    # Create central outlet
    central_outlet = Outlet(
        name="Sagen Baha FPO - Central Office",
        address="Poraiyahat Block, Godda District",
        contact_person="CEO Office",
        is_central=True
    )
    await db.outlets.insert_one(central_outlet.dict())
    
    # Create default products
    default_products = [
        {"name": "Agnistra", "name_hi": "अग्निस्त्र", "unit": "litre", "category": "input"},
        {"name": "Sanjeevini Compost", "name_hi": "संजीविनी खाद", "unit": "kg", "category": "input"},
        {"name": "Multi Seed Extract", "name_hi": "मल्टी सीड एक्स्ट्रेक्ट", "unit": "litre", "category": "input"},
        {"name": "Jeevamrit", "name_hi": "जीवामृत", "unit": "litre", "category": "input"},
        {"name": "Paddy Seeds", "name_hi": "धान बीज", "unit": "kg", "category": "input"},
    ]
    
    for p in default_products:
        product = Product(**p)
        await db.products.insert_one(product.dict())
    
    return {
        "message": "System initialized successfully",
        "admin_username": "admin",
        "admin_password": "admin123",
        "central_outlet_id": central_outlet.id
    }

@api_router.get("/")
async def root():
    return {"message": "FPO Management System API", "version": "1.0"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
